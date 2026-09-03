"""Write human-confirmed desktop decisions to a safe, versioned synthetic workbook.

The writer never edits the supplied workbook.  It treats its yellow input cells
as the only writable surface and uses the existing formulas as the template.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from hashlib import sha256
from pathlib import Path
import json
import os
import re
import shutil
import tempfile
from typing import Iterable
from copy import copy

from openpyxl import load_workbook

from desktop.models import FindingDraft, RiskDecision, ValidationError
from tools.common import DIMS, DOMAINS, assess_all, load_dataset
from tools.export_from_excel import export_workbook, validate_period


RISK_SHEET = "风险登记册"
CONTROL_SHEET = "控制措施表"
CONFIG_SHEET = "参数配置"
FIRST_DATA_ROW = 4
RISK_INPUT_COLUMNS = tuple(range(1, 16)) + (26,)
FORMULA_COLUMNS = tuple(range(16, 26)) + (27, 28)
RISK_ID_RE = re.compile(r"R\d{3}$")
CONTROL_ID_RE = re.compile(r"C\d{3}$")
RISK_HEADERS = ("风险编号", "风险名称", "所属领域", "风险描述", "责任部门", "评估期间")
CONTROL_HEADERS = ("控制编号", "关联风险编号", "评估期间", "控制点描述")


@dataclass(frozen=True, slots=True)
class WorkbookWriteResult:
    workbook_path: Path
    export_dir: Path
    periods: list[str]
    assessed_risks: list[dict]


@dataclass(frozen=True, slots=True)
class _ResolvedDecision:
    decision: RiskDecision
    risk_id: str
    target_row: int | None


def _hash(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _source_path(source: str | Path) -> Path:
    path = Path(source)
    if not path.exists() or not path.is_file() or path.suffix.lower() != ".xlsx":
        raise ValidationError("源工作簿必须是存在的 .xlsx 常规文件")
    return path.resolve()


def _open_checked(path: Path):
    try:
        workbook = load_workbook(path, data_only=False)
        for sheet in (CONFIG_SHEET, RISK_SHEET, CONTROL_SHEET):
            if sheet not in workbook.sheetnames:
                raise KeyError(sheet)
        risks = workbook[RISK_SHEET]
        controls = workbook[CONTROL_SHEET]
        if tuple(_clean_text(risks.cell(3, col).value) for col in range(1, 7)) != RISK_HEADERS:
            raise ValueError("risk headers")
        if tuple(_clean_text(controls.cell(3, col).value) for col in range(1, 5)) != CONTROL_HEADERS:
            raise ValueError("control headers")
        if risks.max_row < FIRST_DATA_ROW or controls.max_row < FIRST_DATA_ROW:
            raise ValueError("reserved rows")
        return workbook
    except (OSError, KeyError, ValueError) as exc:
        raise ValidationError("工作簿结构或表头不符合写入要求") from exc


def _risk_records(ws) -> tuple[list[dict], dict[tuple[str, str], int], set[str]]:
    records: list[dict] = []
    by_key: dict[tuple[str, str], int] = {}
    ids: set[str] = set()
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        risk_id = _clean_text(ws.cell(row, 1).value)
        if not risk_id:
            continue
        period = _clean_text(ws.cell(row, 6).value)
        try:
            period = validate_period(period)
        except ValueError as exc:
            raise ValidationError("工作簿风险编号或期间格式无效") from exc
        if not RISK_ID_RE.fullmatch(risk_id) or _clean_text(ws.cell(row, 3).value) not in DOMAINS:
            raise ValidationError("工作簿风险编号或期间格式无效")
        key = (risk_id, period)
        if key in by_key:
            raise ValidationError("工作簿存在重复的风险编号和期间")
        values = [ws.cell(row, column).value for column in RISK_INPUT_COLUMNS]
        likelihood = values[6]
        dimensions = values[7:15]
        if not isinstance(likelihood, int) or not 1 <= likelihood <= 5 or not any(value is not None for value in dimensions):
            raise ValidationError("工作簿风险录入值无效")
        if any(value is not None and (not isinstance(value, int) or not 1 <= value <= 5) for value in dimensions):
            raise ValidationError("工作簿风险录入值无效")
        records.append({
            "risk_id": risk_id, "name": values[1] or "", "domain": values[2] or "",
            "description": values[3] or "", "owner_dept": values[4] or "",
            "period": period, "likelihood": likelihood,
            **dict(zip(DIMS, dimensions)), "rationale": values[15] or "",
        })
        by_key[key] = row
        ids.add(risk_id)
    return records, by_key, ids


def _control_limit(ws) -> int:
    """Read the existing validation range instead of inserting unstyled rows."""
    limits = []
    for validation in ws.data_validations.dataValidation:
        for match in re.finditer(r"[A-F]4:[A-F](\d+)", str(validation.sqref)):
            limits.append(int(match.group(1)))
    if not limits:
        raise ValidationError("控制措施表缺少预留录入范围")
    return max(limits)


def _control_records(ws, limit: int) -> list[dict]:
    records: list[dict] = []
    ids: set[str] = set()
    for row in range(FIRST_DATA_ROW, limit + 1):
        values = [ws.cell(row, column).value for column in range(1, 7)]
        if all(value is None or not str(value).strip() for value in values):
            continue
        control_id, risk_id, period, description, score, key = values
        control_id = _clean_text(control_id)
        try:
            period = validate_period(_clean_text(period))
        except ValueError as exc:
            raise ValidationError("工作簿控制点录入值无效") from exc
        if (not CONTROL_ID_RE.fullmatch(control_id) or control_id in ids or
                not RISK_ID_RE.fullmatch(_clean_text(risk_id)) or
                not _clean_text(description) or not isinstance(score, int) or not 1 <= score <= 5):
            raise ValidationError("工作簿控制点录入值无效")
        ids.add(control_id)
        records.append({"control_id": control_id, "risk_id": _clean_text(risk_id),
                        "period": period, "description": _clean_text(description),
                        "score": score, "key": "是" if _clean_text(key) in ("是", "1", "true", "True") else "否"})
    return records


def _next_id(existing: Iterable[str], prefix: str) -> str:
    numbers = [int(value[1:]) for value in existing]
    number = (max(numbers) if numbers else 0) + 1
    if number > 999:
        raise ValidationError("工作簿编号容量已满")
    return f"{prefix}{number:03d}"


def _validate_findings(decisions: tuple[RiskDecision, ...], findings: tuple[FindingDraft, ...]) -> None:
    finding_map: dict[str, FindingDraft] = {}
    for item in findings:
        if not isinstance(item, FindingDraft) or item.finding_id in finding_map:
            raise ValidationError("发现项编号必须唯一")
        if item.review_status == "待确认":
            raise ValidationError("存在待确认发现项，不能写入工作簿")
        finding_map[item.finding_id] = item
    referenced: set[str] = set()
    for decision in decisions:
        if not isinstance(decision, RiskDecision):
            raise ValidationError("决策类型无效")
        for finding_id in decision.finding_ids:
            item = finding_map.get(finding_id)
            if item is None or finding_id in referenced:
                raise ValidationError("发现项引用不存在或重复")
            if decision.action == "exclude":
                if item.review_status != "已排除":
                    raise ValidationError("排除决策只能引用已排除发现项")
            elif item.review_status != "已接受":
                raise ValidationError("合并或新建决策只能引用已接受发现项")
            referenced.add(finding_id)
    if any(item.review_status == "已接受" and item.finding_id not in referenced for item in findings):
        raise ValidationError("存在未纳入决策的已接受发现项")


def _decision_values(decision: RiskDecision, risk_id: str) -> list[object]:
    scores = decision.impact_scores or {}
    try:
        period = validate_period(decision.period)
    except ValueError as exc:
        raise ValidationError("评估期间格式无效") from exc
    values = [risk_id, decision.name, decision.domain, decision.description, decision.owner_dept,
              period, decision.likelihood] + [scores.get(dim) for dim in DIMS]
    if not any(value is not None for value in values[7:15]):
        raise ValidationError("至少需要一个影响维度评分")
    return values + [decision.rationale]


def _prepare(source: Path, decisions: Iterable[RiskDecision], findings: Iterable[FindingDraft]) -> dict:
    decisions = tuple(decisions)
    findings = tuple(findings)
    _validate_findings(decisions, findings)
    for decision in decisions:
        if decision.action != "exclude":
            _decision_values(decision, decision.risk_id or "R001")
    workbook = _open_checked(source)
    risks_ws, controls_ws = workbook[RISK_SHEET], workbook[CONTROL_SHEET]
    risks, by_key, risk_ids = _risk_records(risks_ws)
    control_limit = _control_limit(controls_ws)
    controls = _control_records(controls_ws, control_limit)
    if any((control["risk_id"], control["period"]) not in by_key for control in controls):
        raise ValidationError("工作簿控制点找不到对应风险编号和期间")
    resolved: list[_ResolvedDecision] = []
    created_keys: set[tuple[str, str]] = set()
    changed_keys: set[tuple[str, str]] = set()
    next_risk_ids = set(risk_ids)
    blank_rows = [row for row in range(FIRST_DATA_ROW, risks_ws.max_row + 1)
                  if not any(_clean_text(risks_ws.cell(row, column).value) for column in RISK_INPUT_COLUMNS)
                  and isinstance(risks_ws.cell(row, 16).value, str)
                  and risks_ws.cell(row, 16).value.startswith("=")]
    for decision in decisions:
        if decision.action == "exclude":
            continue
        if decision.action == "create":
            expected_risk_id = _next_id(next_risk_ids, "R")
            risk_id = decision.risk_id or expected_risk_id
            if risk_id != expected_risk_id:
                raise ValidationError("新建风险编号必须是下一个连续的 R###")
            key = (risk_id, decision.period or "")
            if key in by_key or key in created_keys or not blank_rows:
                raise ValidationError("风险登记册没有可用的预留行")
            row = blank_rows.pop(0)
            next_risk_ids.add(risk_id)
            created_keys.add(key)
            resolved.append(_ResolvedDecision(decision, risk_id, row))
        else:
            risk_id = decision.risk_id or ""
            key = (risk_id, decision.period or "")
            if not RISK_ID_RE.fullmatch(risk_id) or key not in by_key or key in changed_keys:
                raise ValidationError("未找到唯一匹配的风险编号和期间")
            changed_keys.add(key)
            resolved.append(_ResolvedDecision(decision, risk_id, by_key[key]))
    planned_risks = {(record["risk_id"], record["period"]): dict(record) for record in risks}
    for item in resolved:
        values = _decision_values(item.decision, item.risk_id)
        planned_risks[(item.risk_id, item.decision.period or "")] = {
            "risk_id": values[0], "name": values[1], "domain": values[2], "description": values[3],
            "owner_dept": values[4], "period": values[5], "likelihood": values[6],
            **dict(zip(DIMS, values[7:15])), "rationale": values[15],
        }
    replaced = {(item.risk_id, item.decision.period or "") for item in resolved if item.decision.action == "merge"}
    # IDs remain monotonically allocated even when a merge replaces and removes
    # controls.  Reusing a removed identifier would make an audit trail ambiguous.
    used_control_ids = {record["control_id"] for record in controls}
    remaining_controls = [record for record in controls if (record["risk_id"], record["period"]) not in replaced]
    appended_controls: list[dict] = []
    for item in resolved:
        for control in item.decision.controls:
            control_id = _next_id(used_control_ids, "C")
            used_control_ids.add(control_id)
            appended_controls.append({"control_id": control_id, "risk_id": item.risk_id,
                                      "period": item.decision.period, "description": control.description,
                                      "score": control.score, "key": "是" if control.key else "否"})
    if len(remaining_controls) + len(appended_controls) > control_limit - FIRST_DATA_ROW + 1:
        raise ValidationError("控制措施表没有可用的预留行")
    final_controls = remaining_controls + appended_controls
    warnings = [f"风险 {item.risk_id}/{item.decision.period} 的控制点将完全替换为已确认集合（{len(item.decision.controls)} 条）。"
                for item in resolved if item.decision.action == "merge"]
    preview = {
        "new_risks": [planned_risks[(item.risk_id, item.decision.period or "")] for item in resolved if item.decision.action == "create"],
        "updated_risks": [planned_risks[(item.risk_id, item.decision.period or "")] for item in resolved if item.decision.action == "merge"],
        "new_controls": [{"risk_id": item.risk_id, "period": item.decision.period,
                           "count": len(item.decision.controls), "replacement": item.decision.action == "merge"}
                         for item in resolved],
        "excluded_count": sum(len(item.finding_ids) for item in decisions if item.action == "exclude"),
        "warnings": warnings,
        "assessed_risks": assess_all(list(planned_risks.values()), final_controls,
                                      _config_from_exportable_workbook(workbook)),
    }
    return {"workbook": workbook, "decisions": decisions, "resolved": resolved, "final_controls": final_controls,
            "control_limit": control_limit,
            "findings": findings, "source_hash": _hash(source),
            "preview": preview}


def _config_from_exportable_workbook(workbook) -> dict:
    """Use the one export parser as config authority without touching files."""
    from tools.export_from_excel import read_config
    return read_config(workbook[CONFIG_SHEET])


def _commit_token(prepared: dict) -> str:
    """Hash every resolved user-visible input used by a later commit."""
    payload = {
        "source_sha256": prepared["source_hash"],
        "findings": sorted((item.finding_id, item.review_status) for item in prepared["findings"]),
        "decisions": [{"action": item.action, "finding_ids": list(item.finding_ids)}
                      for item in prepared["decisions"]],
        "resolved": [
            {"action": item.decision.action, "risk_id": item.risk_id, "target_row": item.target_row,
             "inputs": _decision_values(item.decision, item.risk_id),
             "finding_ids": list(item.decision.finding_ids)}
            for item in prepared["resolved"]
        ],
        "controls": prepared["final_controls"],
        "config": _config_from_exportable_workbook(prepared["workbook"]),
    }
    encoded = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")
    return sha256(encoded).hexdigest()


def _preview(source_path: Path, decisions, findings) -> tuple[dict, dict]:
    prepared = _prepare(source_path, decisions, findings)
    preview = dict(prepared["preview"])
    preview["commit_token"] = _commit_token(prepared)
    return preview, prepared


def preview_changes(source, decisions, findings) -> dict:
    """Return deterministic literal-input changes and scores without file mutation."""
    source_path = _source_path(source)
    return _preview(source_path, decisions, findings)[0]


def load_current_controls(source, identities) -> list[dict]:
    """Read only the current controls named by reviewed merge/create identities."""
    source_path = _source_path(source)
    if not isinstance(identities, (list, tuple)):
        raise ValidationError("控制点载入请求必须是列表")
    workbook = _open_checked(source_path)
    try:
        controls = _control_records(workbook[CONTROL_SHEET], _control_limit(workbook[CONTROL_SHEET]))
        loaded = []
        for item in identities:
            if not isinstance(item, dict):
                raise ValidationError("控制点载入标识必须是对象")
            action, risk_id, period, finding_ids = item.get("action"), item.get("risk_id"), item.get("period"), item.get("finding_ids")
            if action not in ("create", "merge") or not isinstance(period, str) or not period:
                raise ValidationError("控制点载入标识无效")
            if not isinstance(finding_ids, (list, tuple)) or not finding_ids or any(not isinstance(value, str) or not value for value in finding_ids):
                raise ValidationError("finding_ids无效")
            if action == "merge" and (not isinstance(risk_id, str) or not RISK_ID_RE.fullmatch(risk_id)):
                raise ValidationError("合并风险编号无效")
            current = [] if action == "create" else [
                {"description": record["description"], "score": record["score"], "key": record["key"] == "是"}
                for record in controls if record["risk_id"] == risk_id and record["period"] == period
            ]
            loaded.append({"finding_ids": list(finding_ids), "action": action,
                           "risk_id": risk_id if action == "merge" else "", "period": period,
                           "controls": current})
        return loaded
    finally:
        workbook.close()


def _write_input_rows(workbook, prepared: dict) -> None:
    risks_ws, controls_ws = workbook[RISK_SHEET], workbook[CONTROL_SHEET]
    for item in prepared["resolved"]:
        values = _decision_values(item.decision, item.risk_id)
        for column, value in zip(RISK_INPUT_COLUMNS, values):
            risks_ws.cell(item.target_row, column).value = value
    control_limit = prepared["control_limit"]
    template_row = next((row for row in range(control_limit, FIRST_DATA_ROW - 1, -1)
                         if controls_ws.cell(row, 1).has_style), None)
    if template_row is None:
        raise ValidationError("控制措施表缺少已设样式的预留行")
    for row in range(FIRST_DATA_ROW, control_limit + 1):
        for column in range(1, 7):
            controls_ws.cell(row, column).value = None
    for row, control in enumerate(prepared["final_controls"], start=FIRST_DATA_ROW):
        # Older synthetic templates reserve validation rows but not their styles.
        # Materialize a style from the last existing input row before writing, so
        # no confirmed control is ever appended to an unstyled row.
        if not controls_ws.cell(row, 1).has_style:
            for column in range(1, 7):
                source = controls_ws.cell(template_row, column)
                target = controls_ws.cell(row, column)
                target._style = copy(source._style)
                target.number_format = source.number_format
                target.alignment = copy(source.alignment)
        for column, key in enumerate(("control_id", "risk_id", "period", "description", "score", "key"), start=1):
            controls_ws.cell(row, column).value = control[key]


def _assert_saved_copy(source: Path, output: Path, prepared: dict) -> None:
    template = _open_checked(source)
    written = _open_checked(output)
    for row in range(FIRST_DATA_ROW, template[RISK_SHEET].max_row + 1):
        for column in FORMULA_COLUMNS:
            if written[RISK_SHEET].cell(row, column).value != template[RISK_SHEET].cell(row, column).value:
                raise ValidationError("公式模板校验失败")
    for item in prepared["resolved"]:
        expected = _decision_values(item.decision, item.risk_id)
        actual = [written[RISK_SHEET].cell(item.target_row, col).value for col in RISK_INPUT_COLUMNS]
        if actual != expected:
            raise ValidationError("工作簿字面输入校验失败")


def write_versioned_workbook(source, decisions, findings, *, expected_commit_token: str,
                             timestamp: str | None = None, output_dir=None) -> WorkbookWriteResult:
    """Build privately, then atomically publish a preview-bound workbook version."""
    source_path = _source_path(source)
    stamp = datetime.now().strftime("%Y%m%d_%H%M") if timestamp is None else timestamp
    if not isinstance(stamp, str) or not re.fullmatch(r"\d{8}_\d{4}", stamp):
        raise ValidationError("timestamp必须为 yyyyMMdd_HHmm")
    parent = (Path(output_dir) if output_dir is not None else source_path.parent).resolve()
    output = (parent / f"audit_risk_register_{stamp}.xlsx").resolve()
    export_dir = output.with_name(f"{output.stem}_data_export")
    if output == source_path:
        raise ValidationError("OUTPUT_EXISTS")

    # This is deliberately before creating the output directory, lock, or temp
    # directory: a stale preview has no filesystem side effects.
    preview, prepared = _preview(source_path, decisions, findings)
    if not isinstance(expected_commit_token, str) or expected_commit_token != preview["commit_token"]:
        raise ValidationError("PREVIEW_STALE")
    source_hash = prepared["source_hash"]

    lock = (parent / f".{output.stem}.lock").resolve()
    lock_created = False
    private_dir: Path | None = None
    published_export = False
    try:
        parent.mkdir(parents=True, exist_ok=True)
        flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
        if hasattr(os, "O_BINARY"):
            flags |= os.O_BINARY
        try:
            descriptor = os.open(lock, flags, 0o600)
        except FileExistsError as exc:
            raise ValidationError("OUTPUT_EXISTS") from exc
        else:
            os.close(descriptor)
            lock_created = True
        if output.exists() or export_dir.exists():
            raise ValidationError("OUTPUT_EXISTS")

        private_dir = Path(tempfile.mkdtemp(prefix=f".{output.stem}.", dir=parent))
        private_workbook = private_dir / output.name
        private_export = private_dir / export_dir.name
        shutil.copy2(source_path, private_workbook)
        workbook = _open_checked(private_workbook)
        _write_input_rows(workbook, prepared)
        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True
            calculation.calcMode = "auto"
        workbook.save(private_workbook)
        _assert_saved_copy(source_path, private_workbook, prepared)
        manifest = export_workbook(private_workbook, private_export)
        assessed: list[dict] = []
        for period in manifest["periods"]:
            config, risks, controls = load_dataset(private_export / period, private_export / "config.json")
            assessed.extend(assess_all(risks, controls, config))
        if _hash(source_path) != source_hash:
            raise ValidationError("PREVIEW_STALE")
        if output.exists() or export_dir.exists():
            raise ValidationError("OUTPUT_EXISTS")
        os.rename(private_export, export_dir)
        published_export = True
        try:
            os.rename(private_workbook, output)  # Publish the workbook last.
        except Exception:
            if published_export and export_dir.exists():
                shutil.rmtree(export_dir)
                published_export = False
            raise
        return WorkbookWriteResult(output, export_dir, manifest["periods"], assessed)
    finally:
        if private_dir is not None and private_dir.exists():
            shutil.rmtree(private_dir)
        if lock_created and lock.exists():
            lock.unlink()
