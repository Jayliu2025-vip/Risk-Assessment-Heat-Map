# -*- coding: utf-8 -*-
"""生成审计风险评估真源工作簿 audit_risk_register.xlsx。

六张表：使用说明 / 参数配置 / 风险登记册 / 控制措施表 / 热力图 / 汇总与优先级。
公式全联动：修改参数配置中的权重、阈值、折减映射后，登记册计算列、
热力图矩阵、汇总排序自动刷新（需 Excel 2019+/365，用到 MINIFS/MAXIFS）。
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import (DEFAULT_CONFIG, DIM_LABELS, DIMS, DOMAINS, LEVEL_COLORS,
                    LEVEL_LABELS, MODEL_VERSION)
from scoring_anchors import load_scoring_anchors
from sample_data import P1, P2, sample_controls, sample_risks

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "audit_risk_register.xlsx")


def argb(hexcolor):
    h = hexcolor.lstrip("#")
    return ("FF" + h) if len(h) == 6 else h


LEVEL_COLORS = {k: argb(v) for k, v in LEVEL_COLORS.items()}

# ---------- 样式常量 ----------
F_TITLE = Font(bold=True, size=14, color="1F4E79")
F_HEAD = Font(bold=True, size=10, color="FFFFFF")
F_SECTION = Font(bold=True, size=11, color="1F4E79")
F_NOTE = Font(size=9, color="808080")
FILL_HEAD = PatternFill("solid", fgColor="1F4E79")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")   # 录入区浅黄
FILL_CALC = PatternFill("solid", fgColor="F2F2F2")    # 公式区浅灰
FILL_SECTION = PatternFill("solid", fgColor="D9E2F1")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

REG_MAX = 203      # 登记册预留行（数据自第 4 行起）
CTRL_MAX = 153     # 控制措施预留行
PRIO_ROWS = 80     # 汇总页优先级清单行数


def style_header(ws, row, cols, height=30):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_band_rules(ws, rng):
    """按阈值给连续数值区域上五档底色（阈值引用参数配置 B14:B17，可随参数变化）。"""
    first = rng.split(":")[0]
    rules = [
        ("extreme", f'AND(ISNUMBER({first}),{first}>=参数配置!$B$14)'),
        ("high", f'AND(ISNUMBER({first}),{first}>=参数配置!$B$15,{first}<参数配置!$B$14)'),
        ("medium", f'AND(ISNUMBER({first}),{first}>=参数配置!$B$16,{first}<参数配置!$B$15)'),
        ("low", f'AND(ISNUMBER({first}),{first}>=参数配置!$B$17,{first}<参数配置!$B$16)'),
        ("minimal", f'AND(ISNUMBER({first}),{first}<参数配置!$B$17)'),
    ]
    for key, formula in rules:
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[formula],
                             fill=PatternFill("solid", fgColor=LEVEL_COLORS[key]),
                             stopIfTrue=True))


def build_config_sheet(wb):
    ws = wb.active
    ws.title = "参数配置"
    ws.sheet_properties.tabColor = "ED7D31"
    ws["A1"] = "参数配置 —— 修改后全簿公式自动联动"
    ws["A1"].font = F_TITLE

    ws["A2"] = "影响维度权重（全领域默认行）"
    ws["A2"].font = F_SECTION
    for i, (d, label) in enumerate(DIM_LABELS.items(), start=3):
        ws[f"A{i}"] = f"{label}权重"
        ws[f"B{i}"] = DEFAULT_CONFIG["weights"][d]
        ws[f"B{i}"].fill = FILL_INPUT
        ws[f"B{i}"].number_format = "0%"
    ws["A11"] = "权重合计（必须=1）"
    ws["B11"] = "=SUM(B3:B10)"
    ws["B11"].number_format = "0%"
    ws.conditional_formatting.add(
        "B11", CellIsRule(operator="notEqual", formula=["1"],
                         fill=PatternFill("solid", fgColor="C00000"), stopIfTrue=True))
    ws["A12"] = "注：权重行之和不为 1 时，合计单元格变红，导出脚本将拒绝执行；维度打分留空=不适用。"
    ws["A12"].font = F_NOTE

    ws["A13"] = "剩余风险等级阈值（≥）"
    ws["A13"].font = F_SECTION
    for i, (label, val) in enumerate([("极高风险", 20), ("高风险", 12),
                                      ("中风险", 6), ("低风险", 3)], start=14):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val
        ws[f"B{i}"].fill = FILL_INPUT

    ws["D3"] = "控制分"
    ws["E3"] = "折减系数"
    style_header(ws, 3, [4, 5], height=20)
    for i, score in enumerate(range(1, 6), start=4):
        ws[f"D{i}"] = score
        ws[f"E{i}"] = DEFAULT_CONFIG["reduction_map"][str(score)]
        ws[f"E{i}"].number_format = "0%"
        ws[f"E{i}"].fill = FILL_INPUT
    ws["D9"] = "短板效应：剩余风险按单条风险“关键控制点最低分”取折减系数（无关键标记则按全部控制点最低分）"
    ws["D9"].font = F_NOTE
    ws["D10"] = "参考折减系数"
    ws["E10"] = DEFAULT_CONFIG["ref_reduction"]
    ws["E10"].number_format = "0%"
    ws["E10"].fill = FILL_INPUT
    ws["D11"] = "仅用于热力图页剩余矩阵的着色基准"
    ws["D11"].font = F_NOTE

    ws["D13"] = "一票否决系数"
    ws["E13"] = DEFAULT_CONFIG["impact_floor_factor"]
    ws["E13"].number_format = "0%"
    ws["E13"].fill = FILL_INPUT
    ws["D14"] = "综合影响 = MAX(加权分, 该系数×最高维度分)：任一维度打高分时致命后果不被其他维度稀释"
    ws["D14"].font = F_NOTE

    ws["A17"] = "颜色图例"
    ws["A17"].font = F_SECTION
    for i, lv in enumerate(["extreme", "high", "medium", "low", "minimal"], start=18):
        cell = ws.cell(row=i, column=1, value=LEVEL_LABELS[lv])
        cell.fill = PatternFill("solid", fgColor=LEVEL_COLORS[lv])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = CENTER

    ws["A24"] = "分领域影响权重矩阵（八维；各行列合计须=1；登记册领域未列出时自动使用“全领域默认”行；维度留空=不适用，权重在已打分维度上重新归一化）"
    ws["A24"].font = F_SECTION
    ws["A25"] = "领域＼维度"
    for j, d in enumerate(DIMS, start=2):
        ws.cell(row=25, column=j, value=DIM_LABELS[d])
    ws.cell(row=25, column=10, value="行合计")
    style_header(ws, 25, range(1, 11), height=20)
    weight_rows = [("全领域默认", DEFAULT_CONFIG["weights"])]
    weight_rows += [(dom, DEFAULT_CONFIG["domain_weights"][dom]) for dom in DOMAINS]
    for i, (dom, wmap) in enumerate(weight_rows, start=26):
        ws.cell(row=i, column=1, value=dom).border = BORDER
        for j, d in enumerate(DIMS, start=2):
            cell = ws.cell(row=i, column=j, value=wmap[d])
            cell.number_format = "0%"
            cell.fill = FILL_INPUT
            cell.alignment = CENTER
            cell.border = BORDER
        sc = ws.cell(row=i, column=10, value=f"=SUM(B{i}:I{i})")
        sc.number_format = "0%"
        sc.alignment = CENTER
        sc.border = BORDER
        ws.conditional_formatting.add(
            f"J{i}", CellIsRule(operator="notEqual", formula=["1"],
                                fill=PatternFill("solid", fgColor="C00000"),
                                stopIfTrue=True))
    set_widths(ws, {"A": 26, "B": 10, "C": 10, "D": 14, "E": 12, "F": 10,
                    "G": 10, "H": 10, "I": 12, "J": 8})
    ws["I1"] = "评估期间清单"
    ws["I1"].font = F_SECTION
    ws["I2"] = "全部"
    ws["I2"].fill = FILL_CALC
    ws["I2"].alignment = CENTER
    ws["I2"].border = BORDER
    for i, p in enumerate([P1, P2], start=3):
        cell = ws.cell(row=i, column=9, value=p)
        cell.fill = FILL_INPUT
        cell.alignment = CENTER
        cell.border = BORDER
    for i in range(5, 9):
        cell = ws.cell(row=i, column=9)
        cell.fill = FILL_INPUT
        cell.border = BORDER
    ws["J3"] = "登记册/控制措施表/热力图的下拉自动引用本清单；新增评估期间在下方空行填写即可"
    ws["J3"].font = F_NOTE
    return ws


def build_anchor_sheet(wb):
    ws = wb.create_sheet("评分锚点", 1)
    ws.sheet_properties.tabColor = "7030A0"
    ws["A1"] = "评分锚点对照表 —— 打分时逐档对照，并在登记册 Z 列填写对应依据"
    ws["A1"].font = F_TITLE
    ws["A2"] = ("金额刻度以国资委资产损失分级为基准，适合央企/国企；非国有企业可按营收规模等比例缩放"
                "（保持相邻档位 5~10 倍跨度），缩放后须全所统一。")
    ws["A2"].font = F_NOTE

    row = 4
    for group in load_scoring_anchors():
        ws.cell(row=row, column=1, value=group["label"]).font = F_SECTION
        row += 1
        for h_i, h in enumerate(["分值", "锚点", "权威依据"], start=1):
            ws.cell(row=row, column=h_i, value=h)
        style_header(ws, row, range(1, 4), height=18)
        row += 1
        for item in group["rows"]:
            a = ws.cell(row=row, column=1, value=item["score"])
            a.alignment = CENTER
            b = ws.cell(row=row, column=2, value=item["anchor"])
            b.alignment = LEFT
            c = ws.cell(row=row, column=3, value=item["source"])
            c.alignment = LEFT
            c.font = F_NOTE
            for cc in (a, b, c):
                cc.border = BORDER
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="权威来源").font = F_SECTION
    sources = [
        "1. 国务院国资委《中央企业违规经营投资责任追究实施办法》（2025 年修订，2026-01-01 施行，替代 37 号令）",
        "2. 国务院令第 493 号《生产安全事故报告和调查处理条例》",
        "3. 《个人信息保护法》第 66 条、《数据安全法》第 45 条、《网络数据安全管理条例》",
        "4. GDPR 第 83 条（2000 万欧元或全球营业额 4%）",
        "5. ACFE《Occupational Fraud 2024: A Report to the Nations》",
        "6. 《监察法》《刑法》职务犯罪管辖标准",
        "7. ISO 31000 / ISO 37301 / ISO 37001 / ISO 22301",
        "8. COSO ERM 2017、IIA《全球内部审计准则》（2024）",
    ]
    for i, s in enumerate(sources, start=1):
        cell = ws.cell(row=row + i, column=1, value=s)
        cell.font = F_NOTE
    set_widths(ws, {"A": 8, "B": 62, "C": 40})
    return ws


def build_register_sheet(wb, risks):
    ws = wb.create_sheet("风险登记册")
    ws.sheet_properties.tabColor = "2E75B6"
    headers = ["风险编号", "风险名称", "所属领域", "风险描述", "责任部门", "评估期间",
               "发生可能性\n(1-5)", "经济损失\n(1-5)", "合规法律\n(1-5)", "运营中断\n(1-5)",
               "声誉舆情\n(1-5)", "舞弊风险\n(1-5)", "战略影响\n(1-5)", "数据安全\n(1-5)",
               "健康安全\n(1-5)",
               "综合影响", "固有风险", "关键控制分", "折减系数", "剩余风险", "剩余等级",
               "影响档\n(辅助)", "固有点\n(辅助)", "建议审计频次", "控制挽回率", "打分依据",
               "优先级排序键\n(辅助)", "选定期间排序键\n(辅助)"]
    ws["A1"] = "风险登记册（黄色列录入，灰色列自动计算，请勿改写公式；影响维度可留空=不适用）"
    ws["A1"].font = F_TITLE
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, range(1, len(headers) + 1))

    for r_off, r in enumerate(risks):
        row = 4 + r_off
        vals = [r["risk_id"], r["name"], r["domain"], r["description"],
                r["owner_dept"], r["period"], r["likelihood"],
                r["imp_financial"], r["imp_compliance"], r["imp_operation"],
                r["imp_reputation"], r["imp_fraud"], r["imp_strategy"],
                r["imp_data"], r["imp_hse"]]
        for c, v in enumerate(vals, start=1):
            if v is None:
                continue
            cell = ws.cell(row=row, column=c, value=v)
            if c >= 7:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
            cell.border = BORDER
            if c <= 6 or (7 <= c <= 15):
                cell.fill = FILL_INPUT
        basis = ws.cell(row=row, column=26, value=r.get("rationale", ""))
        basis.fill = FILL_INPUT
        basis.border = BORDER
        basis.alignment = LEFT

    for row in range(4, REG_MAX + 1):
        wrow = (f'IFERROR(VLOOKUP($C{row},参数配置!$A$26:$I$38,{{2,3,4,5,6,7,8,9}},FALSE),'
                f'VLOOKUP(参数配置!$A$26,参数配置!$A$26:$I$38,{{2,3,4,5,6,7,8,9}},FALSE))')
        formulas = {
            16: (f'=IF($A{row}="","",IF(COUNT(H{row}:O{row})=0,"",'
                 f'ROUND(MAX(SUMPRODUCT(H{row}:O{row},{wrow})/'
                 f'SUMPRODUCT({wrow},--(H{row}:O{row}<>"")),'
                 f'参数配置!$E$13*MAX(H{row}:O{row})),2)))'),
            17: f'=IF($A{row}="","",IF(P{row}="","",ROUND(G{row}*P{row},2)))',
            18: (f'=IF($A{row}="","",IF(COUNTIFS(控制措施表!$B:$B,$A{row},控制措施表!$C:$C,$F{row},'
                 f'控制措施表!$F:$F,"是")=0,IF(COUNTIFS(控制措施表!$B:$B,$A{row},控制措施表!$C:$C,$F{row})=0,"",'
                 f'MINIFS(控制措施表!$E:$E,控制措施表!$B:$B,$A{row},控制措施表!$C:$C,$F{row})),'
                 f'MINIFS(控制措施表!$E:$E,控制措施表!$B:$B,$A{row},控制措施表!$C:$C,$F{row},控制措施表!$F:$F,"是")))'),
            19: f'=IF($A{row}="","",IF(R{row}="",0,VLOOKUP(R{row},参数配置!$D$4:$E$8,2,FALSE)))',
            20: f'=IF($A{row}="","",IF(Q{row}="","",ROUND(Q{row}*(1-S{row}),2)))',
            21: f'=IF($A{row}="","",IF(T{row}="","",IF(T{row}>=参数配置!$B$14,"极高",IF(T{row}>=参数配置!$B$15,"高",IF(T{row}>=参数配置!$B$16,"中",IF(T{row}>=参数配置!$B$17,"低","极低"))))))',
            22: f'=IF($A{row}="","",IF(P{row}="","",ROUND(P{row},0)))',
            23: f'=IF($A{row}="","",IF(Q{row}="","",ROUND(Q{row},0)))',
            24: f'=IF($A{row}="","",IF(U{row}="","",IF(U{row}="极高","每年必审",IF(U{row}="高","每年审计",IF(U{row}="中","两年一轮","按需抽查")))))',
            25: f'=IF($A{row}="","",IF(OR(Q{row}="",Q{row}=0),"",(Q{row}-T{row})/Q{row}))',
            27: (f'=IF($A{row}="","",IF(T{row}="","",'
                 f'T{row}+({REG_MAX + 1}-ROW())/1000000))'),
            28: (f'=IF($A{row}="",NA(),IF(OR('
                 f'\'汇总与优先级\'!$B$2="全部",$F{row}=\'汇总与优先级\'!$B$2),'
                 f'$AA{row},NA()))'),
        }
        for c, fml in formulas.items():
            cell = ws.cell(row=row, column=c, value=fml)
            cell.fill = FILL_CALC
            cell.border = BORDER
            cell.alignment = CENTER
            if c in (16, 17, 20):
                cell.number_format = "0.00"
            if c == 19:
                cell.number_format = "0%"
            if c == 25:
                cell.number_format = "0%"
            if c in (27, 28):
                cell.number_format = "0.000000"

    dv_score = DataValidation(type="whole", operator="between", formula1="1",
                              formula2="5", allow_blank=True,
                              errorTitle="评分超出范围",
                              error="请输入 1~5 的整数；不适用请留空")
    dv_score.add(f"G4:O{REG_MAX}")
    dv_period = DataValidation(type="list", formula1="=参数配置!$I$3:$I$8",
                               allow_blank=True, errorTitle="期间无效",
                               error="请从下拉列表选择（期间清单在参数配置页 I 列维护）")
    dv_period.add(f"F4:F{REG_MAX}")
    dv_domain = DataValidation(type="list", formula1='"' + ",".join(DOMAINS) + '"',
                               allow_blank=True, errorTitle="领域无效", error="请从下拉列表选择")
    dv_domain.add(f"C4:C{REG_MAX}")
    for dv in (dv_score, dv_period, dv_domain):
        ws.add_data_validation(dv)

    # 剩余等级文本上色
    for lv in LEVEL_LABELS.values():
        ws.conditional_formatting.add(
            f"U4:U{REG_MAX}",
            FormulaRule(formula=[f'$U4="{lv}"'],
                        fill=PatternFill("solid", fgColor=LEVEL_COLORS[
                            [k for k, v in LEVEL_LABELS.items() if v == lv][0]]),
                        font=Font(color="FFFFFF", bold=True), stopIfTrue=True))
    # 剩余分值底色随阈值联动
    add_band_rules(ws, f"T4:T{REG_MAX}")

    set_widths(ws, {"A": 9, "B": 20, "C": 12, "D": 42, "E": 11, "F": 10,
                    "G": 9, "H": 9, "I": 9, "J": 9, "K": 9, "L": 9,
                    "M": 9, "N": 9, "O": 9,
                    "P": 9, "Q": 9, "R": 9, "S": 9, "T": 9, "U": 9,
                    "V": 8, "W": 8, "X": 11, "Y": 10, "Z": 40,
                    "AA": 12, "AB": 12})
    ws.column_dimensions["AA"].hidden = True
    ws.column_dimensions["AB"].hidden = True
    ws.freeze_panes = "D4"
    return ws


def build_control_sheet(wb, controls):
    ws = wb.create_sheet("控制措施表")
    ws.sheet_properties.tabColor = "38A3A5"
    headers = ["控制编号", "关联风险编号", "评估期间", "控制点描述", "有效性评分\n(1-5)", "关键控制"]
    ws["A1"] = "控制措施表（每条风险可挂多条控制点；剩余风险按“关键控制最弱分”折减，未标记关键则按全部控制点最弱分）"
    ws["A1"].font = F_TITLE
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, range(1, len(headers) + 1))

    for r_off, c in enumerate(controls):
        row = 4 + r_off
        for ci, v in enumerate([c["control_id"], c["risk_id"], c["period"],
                                c["description"], c["score"], c.get("key", "是")], start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.fill = FILL_INPUT
            cell.border = BORDER
            cell.alignment = LEFT if ci == 4 else CENTER

    dv_score = DataValidation(type="whole", operator="between", formula1="1",
                              formula2="5", allow_blank=True,
                              errorTitle="评分超出范围", error="请输入 1~5 的整数")
    dv_score.add(f"E4:E{CTRL_MAX}")
    dv_rid = DataValidation(type="list", formula1="=风险登记册!$A$4:$A$203",
                            allow_blank=True, errorTitle="编号无效", error="请选择登记册中已有的风险编号")
    dv_rid.add(f"B4:B{CTRL_MAX}")
    dv_period = DataValidation(type="list", formula1="=参数配置!$I$3:$I$8",
                               allow_blank=True, errorTitle="期间无效",
                               error="请从下拉列表选择（期间清单在参数配置页 I 列维护）")
    dv_period.add(f"C4:C{CTRL_MAX}")
    dv_key = DataValidation(type="list", formula1='"是,否"', allow_blank=True,
                            errorTitle="取值无效", error="请选择 是 或 否")
    dv_key.add(f"F4:F{CTRL_MAX}")
    for dv in (dv_score, dv_rid, dv_period, dv_key):
        ws.add_data_validation(dv)

    set_widths(ws, {"A": 10, "B": 13, "C": 10, "D": 52, "E": 11, "F": 9})
    ws.freeze_panes = "A4"
    return ws


def band_fill(l_val, i_val):
    v = l_val * i_val
    cfg_th = [(20, "extreme"), (12, "high"), (6, "medium"), (3, "low")]
    for th, key in cfg_th:
        if v >= th:
            return PatternFill("solid", fgColor=LEVEL_COLORS[key])
    return PatternFill("solid", fgColor=LEVEL_COLORS["minimal"])


def build_heatmap_sheet(wb):
    ws = wb.create_sheet("热力图")
    ws.sheet_properties.tabColor = "C00000"
    ws["A1"] = "风险热力图（可能性 × 综合影响档）"
    ws["A1"].font = F_TITLE
    ws["B3"] = "评估期间筛选："
    ws["C3"] = P2
    ws["C3"].fill = FILL_INPUT
    ws["C3"].font = Font(bold=True)
    dv = DataValidation(type="list", formula1="=参数配置!$I$2:$I$8", allow_blank=False)
    dv.add("C3")
    ws.add_data_validation(dv)

    def grid(top_row, label, count_mode):
        ws.cell(row=top_row, column=2, value=label).font = F_SECTION
        head = top_row + 1
        for j in range(5):
            c = ws.cell(row=head, column=3 + j, value=j + 1)
            c.font = F_HEAD
            c.fill = FILL_HEAD
            c.alignment = CENTER
            c.border = BORDER
        ws.cell(row=head, column=2, value="可能性＼影响档").font = F_HEAD
        ws.cell(row=head, column=2).fill = FILL_HEAD
        ws.cell(row=head, column=2).alignment = CENTER
        crit = f'IF($C$3="全部","<>",$C$3)'
        for i, lik in enumerate(range(5, 0, -1)):
            row = head + 1 + i
            lab = ws.cell(row=row, column=2, value=lik)
            lab.font = Font(bold=True)
            lab.alignment = CENTER
            lab.border = BORDER
            for j in range(5):
                imp = j + 1
                col_letter = get_column_letter(3 + j)
                if count_mode:
                    fml = (f'=COUNTIFS(风险登记册!$F$4:$F${REG_MAX},{crit},'
                           f'风险登记册!$G$4:$G${REG_MAX},$B{row},'
                           f'风险登记册!$V$4:$V${REG_MAX},{col_letter}${head})')
                else:
                    fml = (f'=IFERROR(AVERAGEIFS(风险登记册!$T$4:$T${REG_MAX},'
                           f'风险登记册!$F$4:$F${REG_MAX},{crit},'
                           f'风险登记册!$G$4:$G${REG_MAX},$B{row},'
                           f'风险登记册!$V$4:$V${REG_MAX},{col_letter}${head}),"")')
                cell = ws.cell(row=row, column=3 + j, value=fml)
                cell.alignment = CENTER
                cell.border = BORDER
                cell.number_format = "0" if count_mode else "0.00"
                # 固有矩阵静态底色按格子固有值分档
                if count_mode:
                    cell.fill = band_fill(lik, imp)
        return head + 1

    first = grid(5, "① 固有热力图 —— 格内数字为落入该格的风险数量", True)
    second = grid(13, "② 剩余热力图 —— 格内数字为该位置风险的剩余分均值（底色按均值联动阈值）", False)
    add_band_rules(ws, f"C{second}:G{second + 4}")

    ws.cell(row=21, column=2, value="图例").font = F_SECTION
    for i, lv in enumerate(["extreme", "high", "medium", "low", "minimal"]):
        cell = ws.cell(row=22 + (i // 3), column=3 + (i % 3) * 2, value=LEVEL_LABELS[lv])
        cell.fill = PatternFill("solid", fgColor=LEVEL_COLORS[lv])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = CENTER
    ws.cell(row=25, column=2,
            value="说明：固有矩阵定位＝(发生可能性, 综合影响档)；影响档由精确综合影响四舍五入得到。"
                  "剩余矩阵同一坐标展示控制后的平均剩余分。"
                  "逐条风险的精确剩余分见「汇总与优先级」页。").font = F_NOTE
    set_widths(ws, {"A": 3, "B": 14, "C": 9, "D": 9, "E": 9, "F": 9, "G": 9})
    return ws


def build_summary_sheet(wb):
    ws = wb.create_sheet("汇总与优先级")
    ws.sheet_properties.tabColor = "70AD47"
    ws["A1"] = "汇总统计与年度审计优先级清单"
    ws["A1"].font = F_TITLE
    ws["A2"] = "评估期间（联动热力图页）："
    ws["B2"] = "=热力图!$C$3"
    ws["B2"].font = Font(bold=True)
    crit = 'IF($B$2="全部","<>",$B$2)'

    ws["A4"] = "一、领域汇总"
    ws["A4"].font = F_SECTION
    heads = ["领域", "风险数", "平均剩余分", "最高剩余分", "极高数", "高数", "平均控制挽回率"]
    for i, h in enumerate(heads, start=1):
        ws.cell(row=5, column=i, value=h)
    style_header(ws, 5, range(1, 8), height=22)
    for i, dom in enumerate(DOMAINS, start=6):
        ws.cell(row=i, column=1, value=dom).border = BORDER
        b = ws.cell(row=i, column=2, value=(
            f'=COUNTIFS(风险登记册!$C$4:$C${REG_MAX},$A{i},'
            f'风险登记册!$F$4:$F${REG_MAX},{crit})'))
        c = ws.cell(row=i, column=3, value=(
            f'=IFERROR(ROUND(AVERAGEIFS(风险登记册!$T$4:$T${REG_MAX},'
            f'风险登记册!$C$4:$C${REG_MAX},$A{i},'
            f'风险登记册!$F$4:$F${REG_MAX},{crit}),2),"—")'))
        d = ws.cell(row=i, column=4, value=(
            f'=IFERROR(MAXIFS(风险登记册!$T$4:$T${REG_MAX},'
            f'风险登记册!$C$4:$C${REG_MAX},$A{i},'
            f'风险登记册!$F$4:$F${REG_MAX},{crit}),"—")'))
        e = ws.cell(row=i, column=5, value=(
            f'=COUNTIFS(风险登记册!$U$4:$U${REG_MAX},"极高",'
            f'风险登记册!$C$4:$C${REG_MAX},$A{i},'
            f'风险登记册!$F$4:$F${REG_MAX},{crit})'))
        f_ = ws.cell(row=i, column=6, value=(
            f'=COUNTIFS(风险登记册!$U$4:$U${REG_MAX},"高",'
            f'风险登记册!$C$4:$C${REG_MAX},$A{i},'
            f'风险登记册!$F$4:$F${REG_MAX},{crit})'))
        g = ws.cell(row=i, column=7, value=(
            f'=IFERROR(AVERAGEIFS(风险登记册!$Y$4:$Y${REG_MAX},'
            f'风险登记册!$C$4:$C${REG_MAX},$A{i},'
            f'风险登记册!$F$4:$F${REG_MAX},{crit}),"—")'))
        for cell in (b, c, d, e, f_, g):
            cell.border = BORDER
            cell.alignment = CENTER
            if cell is c:
                cell.number_format = "0.00"
            if cell is g:
                cell.number_format = "0%"
        add_band_rules(ws, f"D{i}")

    prio_title_row = 5 + len(DOMAINS) + 2
    ws.cell(row=prio_title_row, column=1,
            value="二、年度审计优先级清单（按剩余风险降序排名；可用筛选按排名/领域过滤）")
    ws.cell(row=prio_title_row, column=1).font = F_SECTION
    pheads = ["排名", "风险编号", "风险名称", "所属领域", "责任部门", "可能性",
              "固有风险", "关键控制分", "剩余风险", "剩余等级", "建议频次",
              "选源排序键\n(辅助)"]
    prio_head_row = prio_title_row + 1
    for i, h in enumerate(pheads, start=1):
        ws.cell(row=prio_head_row, column=i, value=h)
    style_header(ws, prio_head_row, range(1, len(pheads) + 1), height=24)
    reg_cols = {2: "A", 3: "B", 4: "C", 5: "E", 6: "G",
                7: "Q", 8: "R", 9: "T", 10: "U"}
    for k in range(PRIO_ROWS):
        row = prio_head_row + 1 + k
        selector = ws.cell(row=row, column=12,
                           value=(f'=IFERROR(AGGREGATE(14,6,'
                                  f'风险登记册!$AB$4:$AB${REG_MAX},'
                                  f'ROWS($L${prio_head_row + 1}:L{row})),"")'))
        selector.number_format = "0.000000"
        a = ws.cell(row=row, column=1,
                    value=f'=IF($L{row}="","",ROWS($A${prio_head_row + 1}:A{row}))')
        a.border = BORDER
        a.alignment = CENTER
        for pc, rc in reg_cols.items():
            cell = ws.cell(row=row, column=pc,
                           value=(f'=IF($L{row}="","",INDEX(风险登记册!${rc}$4:${rc}${REG_MAX},'
                                  f'MATCH($L{row},风险登记册!$AA$4:$AA${REG_MAX},0)))'))
            cell.border = BORDER
            cell.alignment = LEFT if pc in (3, 4, 5) else CENTER
            if pc in (7, 9):
                cell.number_format = "0.00"
        kcol = ws.cell(row=row, column=11,
                       value=(f'=IF($L{row}="","",IF($J{row}="极高","每年必审",'
                              f'IF($J{row}="高","每年审计",IF($J{row}="中","两年一轮","按需抽查"))))'))
        kcol.border = BORDER
        kcol.alignment = CENTER
        for lv in LEVEL_LABELS.values():
            key = [kk for kk, vv in LEVEL_LABELS.items() if vv == lv][0]
            ws.conditional_formatting.add(
                f"J{row}:J{row}",
                FormulaRule(formula=[f'$J{row}="{lv}"'],
                            fill=PatternFill("solid", fgColor=LEVEL_COLORS[key]),
                            font=Font(color="FFFFFF", bold=True), stopIfTrue=True))

    set_widths(ws, {"A": 7, "B": 10, "C": 22, "D": 13, "E": 12, "F": 8,
                    "G": 9, "H": 10, "I": 9, "J": 9, "K": 10, "L": 12})
    ws.column_dimensions["L"].hidden = True
    ws.freeze_panes = "A6"
    return ws


def build_readme_sheet(wb):
    ws = wb.create_sheet("使用说明", 0)
    ws.sheet_properties.tabColor = "808080"
    lines = [
        (f"审计风险评估热力图谱 v{MODEL_VERSION} · 使用说明", F_TITLE),
        ("", None),
        ("【模型】综合影响 = MAX(八维加权分, 一票否决系数×最高维度分)：任一维度打 4/5 分时，致命后果不被其他低分维度稀释。", None),
        ("　　风险领域 12 个、4 大类（战略与治理/财务/运营/合规与安全），对齐《企业内部控制应用指引》与《中央企业全面风险管理指引》。", None),
        ("　　八维权重按“所属领域”自动取用分领域权重矩阵（参数配置页 A24 起），未列出的领域用“全领域默认”行。", None),
        ("　　固有风险 = 可能性(1-5) × 综合影响；剩余风险 = 固有 × (1−折减系数)。", None),
        ("　　折减系数由该风险“关键控制点”的最低分查映射表确定（短板效应）；未标记关键控制时退回全部控制点取最低；无控制点则不折减。", None),
        ("　　控制挽回率 = (固有−剩余)/固有，衡量控制环境化解风险的比例，见登记册 Y 列与汇总页领域表。", None),
        ("　　打分原则：各维度只评估其专属后果（如经济损失维度不重复计入舞弊造成的损失），避免维度间重复计分。", None),
        ("　　等级五档：极高≥20 / 高≥12 / 中≥6 / 低≥3 / 极低<3（可在参数配置页调整）。", None),
        ("　　打分锚点见「评分锚点」工作表：国资委资产损失分级、493 号令事故等级、个保法/GDPR 罚则、ACFE 基准等权威对照。", None),
        ("", None),
        ("【日常操作】1. 在「风险登记册」黄色列维护风险条目、可能性与八个影响维度（影响维度可留空，已填项限选 1-5）。", None),
        ("　　　　　　2. 在「控制措施表」为每条风险挂控制点、打有效性分并标记是否关键控制，注意选对评估期间。", None),
        ("　　　　　　3. 打开「热力图」页切换评估期间查看固有/剩余矩阵；「汇总与优先级」页看排序清单与领域挽回率。", None),
        ("　　　　　　4. 调整权重矩阵、阈值、折减映射、一票否决系数请到「参数配置」，全簿公式自动重算。", None),
        ("", None),
        ("【出图出报】运行  python tools/export_from_excel.py  导出 CSV 与 config.json 到 data/export/，", None),
        ("　　　　　　再运行  python tools/generate_report.py  生成 PNG 图集、迁徙矩阵与 PDF 高管简报至 output/。", None),
        ("　　　　　　网页版 web/risk_heatmap.html 可导入同一份 CSV 在浏览器中交互查看与打印汇报。", None),
        ("", None),
        ("【要求】Excel 2019/365 及以上（使用 MINIFS/MAXIFS 函数）；启用宏无需、纯公式实现。", None),
        ("【约定】黄色单元格＝手工录入；灰色单元格＝公式自动计算，请勿覆盖。", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
    ws.column_dimensions["A"].width = 110
    return ws


def main():
    risks, controls = sample_risks(), sample_controls()
    wb = Workbook()
    build_config_sheet(wb)
    build_register_sheet(wb, risks)
    build_control_sheet(wb, controls)
    build_heatmap_sheet(wb)
    build_summary_sheet(wb)
    build_readme_sheet(wb)   # create_sheet(index=0) 自动置于首位
    build_anchor_sheet(wb)   # create_sheet(index=1) 紧随使用说明
    wb.save(XLSX)
    print(f"已生成 {XLSX}  （风险 {len(risks)} 条 / 控制点 {len(controls)} 条 / 两期）")


if __name__ == "__main__":
    main()
