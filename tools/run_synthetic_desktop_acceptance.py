"""Run the local-only synthetic report-to-risk desktop acceptance workflow."""

from __future__ import annotations

import argparse
from dataclasses import asdict
import hashlib
import ipaddress
import logging
import os
from pathlib import Path
import shutil
import socket
import sys
import tempfile
from urllib.parse import urlparse

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from desktop.credentials import CredentialStore
from desktop.extraction import extract_report
from desktop.model_client import ModelClient
from desktop.models import ConfirmedControl, FindingDraft, ModelProfile, RiskDecision
from desktop.ocr import RapidOcrEngine
from desktop.paths import state_db_path, temp_root
from desktop.pipeline import AnalysisPipeline
from desktop.storage import DesktopStore
from desktop.tempfiles import TaskTempFiles
from desktop.workbook_writer import load_current_controls, preview_changes, write_versioned_workbook
from tests.fakes.openai_server import FakeOpenAIServer
from tools.common import DIMS, assess_all, load_dataset
from tools.export_from_excel import export_workbook


FIXTURES = ROOT / "tests" / "fixtures" / "generated"
VERTICAL_REPORT = FIXTURES / "vertical_slice_report.pdf"
SOURCE_WORKBOOK = ROOT / "audit_risk_register.xlsx"
_FULL_BODY_SENTINEL = b"VERTICAL-SYNTHETIC-FULL-BODY-SENTINEL-ONLY"
_SYNTHETIC_KEY = "sk-synthetic"


class AcceptanceError(RuntimeError):
    def __init__(self, code: str) -> None:
        self.code = code
        super().__init__(code)


class MemoryKeyring:
    """A deliberately process-local keyring backend for this synthetic run."""

    def __init__(self) -> None:
        self._values: dict[tuple[str, str], str] = {}

    def set_password(self, service: str, username: str, password: str) -> None:
        self._values[(service, username)] = password

    def get_password(self, service: str, username: str) -> str | None:
        return self._values.get((service, username))

    def delete_password(self, service: str, username: str) -> None:
        self._values.pop((service, username), None)


class OfflineSocketGuard:
    """Process-local network guard for an acceptance run; it never changes a firewall."""

    _LOOPBACK = {"localhost", "127.0.0.1", "::1"}

    def __init__(self) -> None:
        self._create_connection = socket.create_connection
        self._getaddrinfo = socket.getaddrinfo
        self._socket = socket.socket
        self.blocked: list[str] = []

    @classmethod
    def _host(cls, address: object) -> str:
        if not isinstance(address, tuple) or not address or not isinstance(address[0], (str, bytes)):
            raise AcceptanceError("OFFLINE_GUARD")
        host = address[0].decode("ascii", "strict") if isinstance(address[0], bytes) else address[0]
        return host.lower()

    def _allow(self, host: str) -> None:
        try:
            allowed = ipaddress.ip_address(host).is_loopback
        except ValueError:
            allowed = host == "localhost"
        if not allowed:
            self.blocked.append(host)
            raise AcceptanceError("OFFLINE_GUARD")

    def install(self) -> None:
        owner = self

        def guarded_connect(address: object, *args: object, **kwargs: object):
            owner._allow(owner._host(address))
            return owner._create_connection(address, *args, **kwargs)

        def guarded_getaddrinfo(host: str | bytes | None, *args: object, **kwargs: object):
            if host is not None:
                text = host.decode("ascii", "strict") if isinstance(host, bytes) else host
                owner._allow(text.lower())
            return owner._getaddrinfo(host, *args, **kwargs)

        class GuardedSocket(self._socket):
            def connect(self, address: object) -> None:
                owner._allow(owner._host(address))
                return super().connect(address)

            def connect_ex(self, address: object) -> int:
                owner._allow(owner._host(address))
                return super().connect_ex(address)

        socket.create_connection = guarded_connect
        socket.getaddrinfo = guarded_getaddrinfo
        socket.socket = GuardedSocket

    def restore(self) -> None:
        socket.create_connection = self._create_connection
        socket.getaddrinfo = self._getaddrinfo
        socket.socket = self._socket

    def prove_external_blocked(self) -> None:
        try:
            socket.create_connection(("synthetic-external.invalid", 443), timeout=0.01)
        except AcceptanceError as exc:
            if exc.code == "OFFLINE_GUARD" and self.blocked == ["synthetic-external.invalid"]:
                return
        raise AcceptanceError("OFFLINE_GUARD")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _require_exact_loopback(base_url: str) -> None:
    parsed = urlparse(base_url)
    if parsed.hostname not in {"localhost", "127.0.0.1"}:
        raise AcceptanceError("NETWORK_GUARD")


def _inside(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def _build_fixture() -> None:
    from tests.fixtures.build_audit_report_fixtures import build
    build()
    if not VERTICAL_REPORT.is_file():
        raise AcceptanceError("FIXTURE_MISSING")


def _catalog(source: Path, root: Path) -> list[dict]:
    catalog_dir = root / "catalog-export"
    manifest = export_workbook(source, catalog_dir)
    period = manifest["periods"][0]
    _, risks, _ = load_dataset(catalog_dir / period, catalog_dir / "config.json")
    return risks


def _reviewed_findings(pipeline: AnalysisPipeline, store: DesktopStore, task_id: str) -> tuple[FindingDraft, ...]:
    pending = store.list_findings(task_id)
    if len(pending) != 3 or any(item.review_status != "待确认" for item in pending):
        raise AcceptanceError("FINDINGS_INVALID")
    if any("SYNTHETIC TEST DATA" not in item.source_excerpt for item in pending):
        raise AcceptanceError("GROUNDING_INVALID")
    first = FindingDraft(**{**asdict(pending[0]), "title": "Human-edited synthetic finding"})
    pipeline.review_findings(task_id, (first,))
    edited = store.list_findings(task_id)
    if not edited or edited[0].title != "Human-edited synthetic finding" or edited[0].review_status != "待确认":
        raise AcceptanceError("REVIEW_API_INVALID")
    store.set_review_status(task_id, "F-001", "已接受")
    store.set_review_status(task_id, "F-002", "已接受")
    store.set_review_status(task_id, "F-003", "已排除")
    reviewed = tuple(store.list_findings(task_id))
    if [item.review_status for item in reviewed] != ["已接受", "已接受", "已排除"]:
        raise AcceptanceError("REVIEW_STATUS_INVALID")
    return reviewed


def _verify_private_persistence(state_db: Path, report: Path, workbook: Path, output: Path, export_dir: Path) -> None:
    forbidden = (_SYNTHETIC_KEY.encode("utf-8"), _FULL_BODY_SENTINEL,
                 str(report.resolve()).encode("utf-8"), str(workbook.resolve()).encode("utf-8"))
    raw = state_db.read_bytes()
    for value in forbidden:
        if value in raw:
            raise AcceptanceError("PERSISTENCE_PRIVACY")
    for log in state_db.parent.rglob("*.log"):
        body = log.read_bytes()
        if any(value in body for value in forbidden):
            raise AcceptanceError("LOG_PRIVACY")
    for artifact in (output, *(path for path in export_dir.rglob("*") if path.is_file())):
        if any(value in artifact.read_bytes() for value in forbidden):
            raise AcceptanceError("OUTPUT_PRIVACY")


def run_acceptance(root: Path, *, keep_output: bool = False, offline_verify: bool = False) -> dict[str, object]:
    """Execute one deterministic, fully local vertical slice under ``root``."""
    run_root = Path(root).resolve()
    if run_root.exists() and not run_root.is_dir():
        raise AcceptanceError("OUTPUT_ROOT_INVALID")
    run_root.mkdir(parents=True, exist_ok=True)
    _build_fixture()
    report = run_root / "vertical_slice_report.pdf"
    source = run_root / "audit_risk_register.xlsx"
    shutil.copy2(VERTICAL_REPORT, report)
    shutil.copy2(SOURCE_WORKBOOK, source)
    source_sha = _sha256(source)
    old_local_appdata = os.environ.get("LOCALAPPDATA")
    old_logging_disable = logging.root.manager.disable
    logging.disable(logging.INFO)
    os.environ["LOCALAPPDATA"] = str(run_root / "localappdata")
    server: FakeOpenAIServer | None = None
    pipeline: AnalysisPipeline | None = None
    offline_guard = OfflineSocketGuard() if offline_verify else None
    try:
        store = DesktopStore(state_db_path())
        tasks = TaskTempFiles(temp_root())
        catalog = _catalog(source, run_root)
        with FakeOpenAIServer(mode="vertical") as running_server:
            server = running_server
            if offline_guard is not None:
                offline_guard.install()
                offline_guard.prove_external_blocked()
            _require_exact_loopback(server.base_url)
            profile = ModelProfile("synthetic-local", server.base_url, "synthetic-model", False)
            memory = MemoryKeyring()
            credentials = CredentialStore(memory)
            credentials.set_api_key(profile.name, _SYNTHETIC_KEY)

            def make_client(checked_profile: ModelProfile, api_key: str) -> ModelClient:
                _require_exact_loopback(checked_profile.base_url)
                return ModelClient(checked_profile, api_key)

            extracted = []

            def extractor(path: Path, task_dir: Path):
                result = extract_report(path, task_dir, RapidOcrEngine())
                extracted.append(result)
                return result

            pipeline = AnalysisPipeline(
                store, tasks, extractor,
                make_client, lambda _: profile, credentials.get_api_key, catalog,
            )
            task = pipeline.start(report, profile.name)
            complete = pipeline.wait(task.task_id, timeout=120)
            if complete.status != "待复核":
                raise AcceptanceError("PIPELINE_FAILED")
            third_page = next((block for result in extracted for block in result.blocks if block.locator == "第 3 页"), None)
            normalized_ocr = "" if third_page is None else "".join(third_page.text.upper().split())
            if (third_page is None or third_page.method != "ocr" or "SYNTHETICTESTDATA" not in normalized_ocr
                    or "LOCATORGAMMA" not in normalized_ocr):
                raise AcceptanceError("OCR_NOT_PROVEN")
            findings = _reviewed_findings(pipeline, store, task.task_id)
            if [item.finding_id for item in findings] != ["F-001", "F-002", "F-003"]:
                raise AcceptanceError("FINDING_IDS_INVALID")
            decision = RiskDecision(
                action="create", finding_ids=("F-001", "F-002"), risk_id="",
                name="Synthetic merged risk", domain="资金活动",
                description="Fictional combined finding from local test data.", owner_dept="Synthetic Audit",
                period="2026H2", likelihood=3, impact_scores={dimension: 2 for dimension in DIMS},
                rationale="Human-confirmed synthetic evidence.",
                controls=(ConfirmedControl("Confirmed synthetic current control", 4, True),),
            )
            excluded = RiskDecision(action="exclude", finding_ids=("F-003",))
            current = load_current_controls(source, [{"action": "create", "risk_id": "", "period": "2026H2", "finding_ids": ["F-001", "F-002"]}])
            if current != [{"finding_ids": ["F-001", "F-002"], "action": "create", "risk_id": "", "period": "2026H2", "controls": []}]:
                raise AcceptanceError("CURRENT_CONTROLS_INVALID")
            preview = preview_changes(source, (decision, excluded), findings)
            result = write_versioned_workbook(source, (decision, excluded), findings,
                                              expected_commit_token=preview["commit_token"], timestamp="20260903_1111",
                                              output_dir=run_root / "versions")
            if result.periods != sorted(set(result.periods)) or "2026H2" not in result.periods:
                raise AcceptanceError("PERIOD_INVALID")
            output = result.workbook_path
            if not output.is_file() or not _inside(output, run_root) or not _inside(result.export_dir, run_root):
                raise AcceptanceError("OUTPUT_LOCATION_INVALID")
            workbook = load_workbook(output, data_only=False)
            try:
                register = workbook["风险登记册"]
                if not any(register.cell(row, 1).value == "R025" and register.cell(row, 6).value == "2026H2"
                           for row in range(4, register.max_row + 1)):
                    raise AcceptanceError("WORKBOOK_INVALID")
            finally:
                workbook.close()
            config, risks, controls = load_dataset(result.export_dir / "2026H2", result.export_dir / "config.json")
            expected = assess_all(risks, controls, config)
            actual = [item for item in result.assessed_risks if item["period"] == "2026H2"]
            if actual != expected:
                raise AcceptanceError("RESIDUAL_MISMATCH")
            if _sha256(source) != source_sha:
                raise AcceptanceError("SOURCE_CHANGED")
            if tasks.task_dir(task.task_id).exists():
                raise AcceptanceError("TEMP_NOT_CLEAN")
            if not server.requests or any(request.get("_remote_host") not in {"127.0.0.1", "::1"} for request in server.requests):
                raise AcceptanceError("NETWORK_GUARD")
            if any("authorization" in str(request).lower() for request in server.requests):
                raise AcceptanceError("AUTHORIZATION_RECORDED")
            _verify_private_persistence(state_db_path(), report, source, output, result.export_dir)
            return {"findings": 3, "accepted": 2, "excluded": 1, "period": "2026H2", "periods": result.periods,
                    "source_unchanged": True, "temp_clean": True,
                    "workbook": output, "export_dir": result.export_dir, "residual": actual,
                    "source": source, "source_sha256": source_sha, "post_source_sha256": _sha256(source),
                    "task_temp_dir": tasks.task_dir(task.task_id), "state_db": state_db_path(),
                    "server_requests": tuple(server.requests), "report": report,
                    "ocr": {"locator": third_page.locator, "method": third_page.method},
                    "offline_guard": offline_guard is not None and offline_guard.blocked == ["synthetic-external.invalid"],
                    "output_files": tuple([output, *(path for path in result.export_dir.rglob("*") if path.is_file())])}
    finally:
        if pipeline is not None:
            pipeline.close()
        if offline_guard is not None:
            offline_guard.restore()
        if old_local_appdata is None:
            os.environ.pop("LOCALAPPDATA", None)
        else:
            os.environ["LOCALAPPDATA"] = old_local_appdata
        logging.disable(old_logging_disable)


def _keep_root(value: str) -> Path:
    requested = Path(value).expanduser()
    if requested.exists():
        if not requested.is_dir():
            raise AcceptanceError("OUTPUT_ROOT_INVALID")
        return requested.resolve()
    parent = requested.parent.resolve()
    if not parent.is_dir() or requested.name in {"", ".", ".."}:
        raise AcceptanceError("OUTPUT_ROOT_INVALID")
    return requested.resolve()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(add_help=True)
    parser.add_argument("--keep-output", metavar="DIR")
    parser.add_argument("--offline-verify", action="store_true")
    args = parser.parse_args(argv)
    try:
        if args.keep_output:
            root = _keep_root(args.keep_output)
            if root.exists():
                root = Path(tempfile.mkdtemp(prefix="synthetic-desktop-acceptance-", dir=root))
            result = run_acceptance(root, keep_output=True, offline_verify=args.offline_verify)
        else:
            with tempfile.TemporaryDirectory(prefix="rahm-desktop-acceptance-") as directory:
                result = run_acceptance(Path(directory), offline_verify=args.offline_verify)
        if args.offline_verify and result["offline_guard"] is True:
            print("OFFLINE_GUARD_OK loopback_only=true")
        print("DESKTOP_ACCEPTANCE_OK findings={findings} accepted={accepted} excluded={excluded} period={period} source_unchanged=true temp_clean=true".format(**result))
        return 0
    except AcceptanceError as exc:
        print(f"DESKTOP_ACCEPTANCE_FAILED code={exc.code}", file=sys.stderr)
        return 1
    except Exception:
        print("DESKTOP_ACCEPTANCE_FAILED code=UNEXPECTED", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
