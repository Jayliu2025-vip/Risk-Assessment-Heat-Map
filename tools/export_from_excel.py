# -*- coding: utf-8 -*-
"""从真源工作簿 audit_risk_register.xlsx 一键导出：
    data/export/{期间}/risks.csv + controls.csv
    data/export/config.json（权重/阈值/折减映射随参数配置页同步）

只读取录入列（字面值），不依赖公式计算结果；评分模型计算由
generate_report.py 与 risk_heatmap.html 各自按同一公式实现。
"""
import argparse
import csv
import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import CONTROL_FIELDS, DIMS, MODEL_VERSION, RISK_FIELDS

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_XLSX = os.path.join(ROOT, "audit_risk_register.xlsx")
OUT_DIR = os.path.join(ROOT, "data", "export")


class ExportWorkbookError(ValueError):
    """A workbook cannot be exported without violating its input contract."""


def read_config(ws):
    weights = {
        "imp_financial": float(ws["B3"].value),
        "imp_compliance": float(ws["B4"].value),
        "imp_operation": float(ws["B5"].value),
        "imp_reputation": float(ws["B6"].value),
        "imp_fraud": float(ws["B7"].value),
        "imp_strategy": float(ws["B8"].value),
        "imp_data": float(ws["B9"].value),
        "imp_hse": float(ws["B10"].value),
    }
    reduction_map = {}
    for row in range(4, 9):
        score = ws[f"D{row}"].value
        if score is not None:
            reduction_map[str(int(score))] = round(float(ws[f"E{row}"].value), 4)
    thresholds = {"extreme": int(ws["B14"].value), "high": int(ws["B15"].value),
                  "medium": int(ws["B16"].value), "low": int(ws["B17"].value)}
    ref_reduction = float(ws["E10"].value or 0.4)
    impact_floor = float(ws["E13"].value or 0.75)
    dim_keys = list(DIMS)
    domain_weights = {}
    for row in range(27, 39):
        dom = ws[f"A{row}"].value
        if not dom:
            continue
        domain_weights[str(dom).strip()] = {
            k: round(float(ws.cell(row=row, column=2 + j).value or 0), 4)
            for j, k in enumerate(dim_keys)
        }
    return {"version": MODEL_VERSION, "weights": weights,
            "domain_weights": domain_weights,
            "impact_floor_factor": impact_floor,
            "reduction_map": reduction_map,
            "thresholds": thresholds, "ref_reduction": ref_reduction}


def read_rows(ws, max_col, start=4):
    rows = []
    r = start
    while True:
        first = ws.cell(row=r, column=1).value
        if first is None or str(first).strip() == "":
            break
        rows.append([ws.cell(row=r, column=c).value for c in range(1, max_col + 1)])
        r += 1
    return rows


def export_workbook(xlsx_path, out_dir=OUT_DIR):
    """Export literal input columns from a workbook and return its manifest.

    This API intentionally does not terminate the process.  Desktop callers can
    therefore clean up an incomplete versioned copy while the command line keeps
    its established Chinese diagnostics in :func:`main`.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=False)
        config_ws = wb["参数配置"]
        risks_ws = wb["风险登记册"]
        controls_ws = wb["控制措施表"]
    except (OSError, KeyError, ValueError) as exc:
        raise ExportWorkbookError("[错误] 工作簿结构不符合导出要求。") from exc
    cfg = read_config(config_ws)
    all_rows = {"全领域默认": cfg["weights"]}
    all_rows.update(cfg["domain_weights"])
    for name, w in all_rows.items():
        wsum = round(sum(w.values()), 6)
        if abs(wsum - 1.0) > 1e-6:
            raise ExportWorkbookError(f"[错误] 权重行[{name}]之和为 {wsum}，必须等于 1。请先修正「参数配置」页。")

    risks, controls = [], []
    for vals in read_rows(risks_ws, max_col=26):
        rid, name, dom, desc, dept, period, lik = vals[:7]
        dims = [None if v is None or str(v).strip() == "" else int(v)
                for v in vals[7:15]]
        scored = [d for d in dims if d is not None]
        if not scored:
            raise ExportWorkbookError(f"[错误] 风险 {rid} 至少需要为一个影响维度打分（其余可留空=不适用）。")
        for label, v in zip(["可能性"] + [f"影响-{d}" for d in DIMS],
                            [lik] + dims):
            if v is None and label != "可能性":
                continue
            if not (isinstance(v, int) and 1 <= v <= 5):
                raise ExportWorkbookError(f"[错误] 风险 {rid} 的打分 {label}={v} 超出 1~5。")
        risks.append({"risk_id": str(rid).strip(), "name": name, "domain": dom,
                      "description": desc or "", "owner_dept": dept or "",
                      "period": period, "likelihood": int(lik),
                      "rationale": vals[25] or "",
                      **dict(zip(DIMS, dims))})
    for vals in read_rows(controls_ws, max_col=6):
        cid, rid, period, desc, score, key = vals
        controls.append({"control_id": str(cid).strip(), "risk_id": str(rid).strip(),
                         "period": period, "description": desc or "",
                         "score": int(score),
                         "key": "是" if str(key).strip() in ("是", "1", "true", "True") else "否"})

    periods = sorted({r["period"] for r in risks})
    output = Path(out_dir)
    paths = []
    for p in periods:
        d = output / str(p)
        d.mkdir(parents=True, exist_ok=True)
        with (d / "risks.csv").open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=RISK_FIELDS)
            w.writeheader()
            w.writerows([r for r in risks if r["period"] == p])
        with (d / "controls.csv").open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=CONTROL_FIELDS)
            w.writeheader()
            w.writerows([c for c in controls if c["period"] == p])
        paths.extend((d / "risks.csv", d / "controls.csv"))
    config_path = output / "config.json"
    with config_path.open("w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    paths.append(config_path)
    orphan = [c["control_id"] for c in controls
              if not any(r["risk_id"] == c["risk_id"] and r["period"] == c["period"]
                         for r in risks)]
    return {"risks": len(risks), "controls": len(controls), "periods": periods,
            "paths": [str(path) for path in paths], "orphan_controls": orphan}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    args = ap.parse_args()
    try:
        result = export_workbook(args.xlsx, OUT_DIR)
    except ExportWorkbookError as exc:
        print(str(exc), file=sys.stderr)
        return 1
    except (OSError, ValueError, TypeError, KeyError):
        print("[错误] 工作簿结构不符合导出要求。", file=sys.stderr)
        return 1
    print(f"导出完成：{result['risks']} 条风险 / {result['controls']} 个控制点 / {len(result['periods'])} 期 "
          f"({', '.join(map(str, result['periods']))}) -> {os.path.relpath(OUT_DIR, ROOT)}")
    if result["orphan_controls"]:
        print(f"[警告] {len(result['orphan_controls'])} 个控制点找不到对应（风险编号+期间）：{result['orphan_controls']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
