"""End-to-end contract for the local, synthetic desktop workflow."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
import socket
import sqlite3
import zipfile

from openpyxl import load_workbook
from openpyxl import Workbook

from tools.common import assess_all, load_dataset


class AlwaysFailingOcr:
    def read(self, image_path: Path) -> str:
        from desktop.ocr import OcrError
        raise OcrError("OCR_FAILED", "synthetic test failure")


class DesktopVerticalSliceTests(unittest.TestCase):
    def test_synthetic_desktop_acceptance_returns_a_private_verified_result(self) -> None:
        from tools.run_synthetic_desktop_acceptance import run_acceptance

        with tempfile.TemporaryDirectory() as temporary:
            result = run_acceptance(Path(temporary), keep_output=True)
            self.assertEqual(result["findings"], 3)
            self.assertEqual(result["accepted"], 2)
            self.assertEqual(result["excluded"], 1)
            self.assertEqual(result["period"], "2026H2")
            self.assertIn("2026H2", result["periods"])
            workbook = load_workbook(result["workbook"], data_only=False)
            try:
                self.assertTrue(any(workbook["风险登记册"].cell(row, 1).value == "R025"
                                    for row in range(4, workbook["风险登记册"].max_row + 1)))
            finally:
                workbook.close()
            config, risks, controls = load_dataset(result["export_dir"] / "2026H2", result["export_dir"] / "config.json")
            self.assertEqual(result["residual"], assess_all(risks, controls, config))
            self.assertTrue(result["source_unchanged"])
            self.assertTrue(result["temp_clean"])
            self.assertEqual(result["source_sha256"], result["post_source_sha256"])
            self.assertFalse(Path(result["task_temp_dir"]).exists())
            self.assertEqual(result["ocr"], {"locator": "第 3 页", "method": "ocr"})
            self.assertTrue(all(request["_remote_host"] in {"127.0.0.1", "::1"} for request in result["server_requests"]))
            self.assertNotIn("authorization", str(result["server_requests"]).lower())
            raw_db = Path(result["state_db"]).read_bytes()
            for forbidden in (b"sk-synthetic", b"VERTICAL-SYNTHETIC-FULL-BODY-SENTINEL-ONLY",
                              str(Path(result["report"]).resolve()).encode("utf-8"),
                              str(Path(result["source"]).resolve()).encode("utf-8")):
                self.assertNotIn(forbidden, raw_db)
            report_path = str(Path(result["report"]).resolve()).encode("utf-8")
            self.assertTrue(all(report_path not in Path(path).read_bytes() for path in result["output_files"]))

    def test_acceptance_rejects_ocr_vision_fallback(self) -> None:
        from tools.run_synthetic_desktop_acceptance import AcceptanceError, run_acceptance

        with tempfile.TemporaryDirectory() as temporary:
            with patch("tools.run_synthetic_desktop_acceptance.RapidOcrEngine", return_value=AlwaysFailingOcr()):
                with self.assertRaisesRegex(AcceptanceError, "OCR_NOT_PROVEN"):
                    run_acceptance(Path(temporary), keep_output=True)

    def test_acceptance_rejects_broken_pending_review_api(self) -> None:
        from desktop.pipeline import AnalysisPipeline
        from tools.run_synthetic_desktop_acceptance import AcceptanceError, run_acceptance

        with tempfile.TemporaryDirectory() as temporary:
            with patch.object(AnalysisPipeline, "review_findings", return_value=[]):
                with self.assertRaisesRegex(AcceptanceError, "REVIEW_API_INVALID"):
                    run_acceptance(Path(temporary), keep_output=True)

    def test_offline_mode_blocks_external_socket_before_any_request(self) -> None:
        from tools.run_synthetic_desktop_acceptance import run_acceptance

        with tempfile.TemporaryDirectory() as temporary:
            result = run_acceptance(Path(temporary), keep_output=True, offline_verify=True)
        self.assertTrue(result["offline_guard"])

    def test_offline_guard_blocks_direct_socket_paths_but_allows_loopback_model_client(self) -> None:
        from desktop.model_client import ModelClient
        from desktop.models import ModelProfile
        from tests.fakes.openai_server import FakeOpenAIServer
        from tools.run_synthetic_desktop_acceptance import AcceptanceError, OfflineSocketGuard

        guard = OfflineSocketGuard()
        guard.install()
        try:
            with self.assertRaisesRegex(AcceptanceError, "OFFLINE_GUARD"):
                socket.create_connection(("203.0.113.1", 443), timeout=0.01)
            direct = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            direct.settimeout(0.01)
            try:
                with self.assertRaisesRegex(AcceptanceError, "OFFLINE_GUARD"):
                    direct.connect(("203.0.113.1", 443))
                with self.assertRaisesRegex(AcceptanceError, "OFFLINE_GUARD"):
                    direct.connect_ex(("203.0.113.1", 443))
            finally:
                direct.close()
            with self.assertRaisesRegex(AcceptanceError, "OFFLINE_GUARD"):
                socket.getaddrinfo("synthetic-external.invalid", 443)
            with FakeOpenAIServer(content="OK") as server:
                profile = ModelProfile("loopback", server.base_url, "synthetic-model", False)
                with ModelClient(profile, "sk-synthetic") as client:
                    self.assertTrue(client.test_connection())
        finally:
            guard.restore()
        self.assertTrue(guard.blocked)

    def test_vertical_fake_rejects_redacted_transported_evidence(self) -> None:
        from desktop import model_client
        from tools.run_synthetic_desktop_acceptance import AcceptanceError, run_acceptance

        original = model_client.build_analysis_messages

        def redact(*args, **kwargs):
            messages = original(*args, **kwargs)
            messages[-1] = {"role": "user", "content": "REDACTED"}
            return messages

        with tempfile.TemporaryDirectory() as temporary:
            with patch("desktop.model_client.build_analysis_messages", side_effect=redact):
                with self.assertRaisesRegex(AcceptanceError, "PIPELINE_FAILED"):
                    run_acceptance(Path(temporary), keep_output=True)

    def test_workbook_privacy_rejects_hidden_compressed_cell_and_relationship_path(self) -> None:
        from tools.run_synthetic_desktop_acceptance import AcceptanceError, _verify_private_persistence

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            report = root / "vertical_slice_report.pdf"
            source = root / "audit_risk_register.xlsx"
            output = root / "version.xlsx"
            export = root / "export"
            report.write_bytes(b"synthetic")
            source.write_bytes(b"synthetic")
            sqlite3.connect(root / "state.db").close()
            workbook = Workbook()
            workbook.active["A1"] = (str(report.resolve()) + " ") * 160
            workbook.save(output)
            with zipfile.ZipFile(output, "a", zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("xl/externalLinks/_rels/externalLink1.xml.rels",
                                 ("<Relationships><Relationship Target=\"" + str(report.resolve()) * 160
                                  + "\"/></Relationships>").encode("utf-8"))
            self.assertNotIn(str(report.resolve()).encode("utf-8"), output.read_bytes())
            export.mkdir()
            with self.assertRaisesRegex(AcceptanceError, "OUTPUT_PRIVACY"):
                _verify_private_persistence(root / "state.db", report, source, output, export)


if __name__ == "__main__":
    unittest.main()
