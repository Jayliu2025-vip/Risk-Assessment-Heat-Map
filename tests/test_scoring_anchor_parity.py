# -*- coding: utf-8 -*-
"""评分锚点的规范数据、浏览器副本和工作簿展示必须完全一致。"""

import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.common import DIMS


ROOT = Path(__file__).resolve().parents[1]
EXPECTED_KEYS = ["likelihood", *DIMS]


def load_json_groups():
    path = ROOT / "data" / "scoring_anchors.json"
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


class ScoringAnchorParityTests(unittest.TestCase):
    def test_canonical_json_has_all_nine_complete_groups(self):
        groups = load_json_groups()
        self.assertIsNotNone(groups, "必须提供 data/scoring_anchors.json 作为评分锚点真源")
        self.assertEqual([group["key"] for group in groups], EXPECTED_KEYS)
        self.assertEqual(len(groups), 9)
        self.assertEqual(sum(len(group["rows"]) for group in groups), 45)
        for group in groups:
            with self.subTest(key=group["key"]):
                self.assertTrue(group["label"].strip())
                self.assertEqual([row["score"] for row in group["rows"]], [1, 2, 3, 4, 5])
                for row in group["rows"]:
                    self.assertTrue(row["anchor"].strip())
                    self.assertTrue(row["source"].strip())

    def test_browser_copy_roundtrips_exactly_to_canonical_json(self):
        groups = load_json_groups()
        self.assertIsNotNone(groups)
        script = (ROOT / "web" / "scoring_anchors.js").read_text(encoding="utf-8").strip()
        prefix = "window.SCORING_ANCHORS = "
        self.assertTrue(script.startswith(prefix))
        self.assertEqual(json.loads(script[len(prefix):].removesuffix(";")), groups)

    def test_web_uses_standalone_canonical_anchors_instead_of_inline_copy(self):
        html = (ROOT / "web" / "risk_heatmap.html").read_text(encoding="utf-8")
        self.assertIn('<script src="scoring_anchors.js"></script>', html)
        self.assertIn("SCORING_ANCHORS.map", html)
        self.assertNotIn("const ANCHORS = [", html)

    def test_workbook_has_nine_ordered_anchor_blocks_with_five_score_rows_each(self):
        groups = load_json_groups()
        self.assertIsNotNone(groups)
        workbook = load_workbook(ROOT / "audit_risk_register.xlsx", data_only=False)
        sheet = workbook["评分锚点"]
        row = 4
        for group in groups:
            with self.subTest(key=group["key"]):
                self.assertEqual(sheet.cell(row=row, column=1).value, group["label"])
                self.assertEqual(
                    [sheet.cell(row=row + 2 + offset, column=1).value for offset in range(5)],
                    [1, 2, 3, 4, 5],
                )
                self.assertEqual(
                    [sheet.cell(row=row + 2 + offset, column=2).value for offset in range(5)],
                    [item["anchor"] for item in group["rows"]],
                )
                self.assertEqual(
                    [sheet.cell(row=row + 2 + offset, column=3).value for offset in range(5)],
                    [item["source"] for item in group["rows"]],
                )
            row += 8


if __name__ == "__main__":
    unittest.main()
