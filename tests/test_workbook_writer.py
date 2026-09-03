"""Synthetic-only contracts for writing reviewed findings into a versioned workbook."""

from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
import threading
import unittest
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from desktop.models import ConfirmedControl, FindingDraft, RiskDecision, ValidationError
from desktop.workbook_writer import (WorkbookWriteResult, preview_changes,
                                     write_versioned_workbook)
from tools.common import DIMS, assess_all, load_dataset
from tools.export_from_excel import export_workbook


ROOT = Path(__file__).resolve().parents[1]
SYNTHETIC_SOURCE = ROOT / "audit_risk_register.xlsx"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def finding(fid: str, status: str = "已接受") -> FindingDraft:
    return FindingDraft(
        task_id="synthetic-task", finding_id=fid, title=f"问题 {fid}",
        fact_summary="合成事实", source_page="1", source_excerpt="合成摘录",
        matched_risk_id="", domain="采购与外包", likelihood=4,
        impact_scores={dim: 3 for dim in DIMS}, rationale="合成依据",
        needs_review=False, review_status=status,
    )


def create_decision(fid: str = "F-create", risk_id: str = "") -> RiskDecision:
    return RiskDecision(
        action="create", finding_ids=(fid,), risk_id=risk_id, name="新增合成风险",
        domain="采购与外包", description="新增风险描述", owner_dept="审计部",
        period="2026H2", likelihood=4, impact_scores={dim: 3 for dim in DIMS},
        rationale="确认后的合成依据", controls=(ConfirmedControl("双人复核", 4, True),),
    )


def merge_decision(fid: str = "F-merge", controls: tuple[ConfirmedControl, ...] | None = None) -> RiskDecision:
    return RiskDecision(
        action="merge", finding_ids=(fid,), risk_id="R001", name="更新后的合成风险",
        domain="采购与外包", description="更新风险描述", owner_dept="审计部",
        period="2026H1", likelihood=5, impact_scores={dim: 4 for dim in DIMS},
        rationale="更新后的合成依据", controls=(ConfirmedControl("更新控制", 5, True),)
        if controls is None else controls,
    )


class WorkbookWriterTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.source = self.root / "synthetic_source.xlsx"
        shutil.copy2(SYNTHETIC_SOURCE, self.source)
        self.source_hash = sha256(self.source)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def assert_source_unchanged(self) -> None:
        self.assertEqual(sha256(self.source), self.source_hash)

    @staticmethod
    def commit(source, decisions, findings, **kwargs) -> WorkbookWriteResult:
        token = preview_changes(source, decisions, findings)["commit_token"]
        return write_versioned_workbook(source, decisions, findings, expected_commit_token=token, **kwargs)

    def test_create_blank_id_merge_and_exclude_write_versioned_copy(self) -> None:
        result = self.commit(
            self.source,
            (create_decision(), merge_decision(), RiskDecision(action="exclude", finding_ids=("F-exclude",))),
            (finding("F-create"), finding("F-merge"), finding("F-exclude", "已排除")),
            timestamp="20260903_1200", output_dir=self.root / "versions",
        )
        self.assertIsInstance(result, WorkbookWriteResult)
        self.assertEqual(result.workbook_path.name, "audit_risk_register_20260903_1200.xlsx")
        self.assertTrue(result.workbook_path.is_file())
        self.assertIn("2026H2", result.periods)
        self.assertEqual(result.assessed_risks, self._all_assessed(result.export_dir, result.periods))
        wb = load_workbook(result.workbook_path, data_only=False)
        register = wb["风险登记册"]
        created = next(row for row in range(4, 204) if register.cell(row, 1).value == "R025")
        self.assertEqual([register.cell(created, c).value for c in range(1, 16)],
                         ["R025", "新增合成风险", "采购与外包", "新增风险描述", "审计部", "2026H2", 4] + [3] * 8)
        merged = next(row for row in range(4, 204) if register.cell(row, 1).value == "R001" and register.cell(row, 6).value == "2026H1")
        self.assertEqual(register.cell(merged, 2).value, "更新后的合成风险")
        self.assert_source_unchanged()

    def test_preview_is_deterministic_and_does_not_mutate(self) -> None:
        before = sha256(self.source)
        args = (self.source, (create_decision(),), (finding("F-create"),))
        first = preview_changes(*args)
        self.assertEqual(first, preview_changes(*args))
        self.assertEqual(first["new_risks"][0]["risk_id"], "R025")
        self.assertEqual(first["excluded_count"], 0)
        self.assertIn("assessed_risks", first)
        self.assertEqual(sha256(self.source), before)
        self.assertEqual(list(self.root.glob("*.xlsx")), [self.source])

    def test_periods_reject_path_syntax_without_escaping_export_root(self) -> None:
        unsafe_periods = ("../escaped_period", ".", "..", "2026/H2", "2026\\H2", str(self.root / "absolute"))
        for period in unsafe_periods:
            with self.subTest(period=period):
                decision = create_decision()
                decision.period = period
                with self.assertRaises((ValidationError, ValueError)):
                    preview_changes(self.source, (decision,), (finding("F-create"),))
        for period in unsafe_periods:
            source = self.root / f"period-{len(period)}.xlsx"
            shutil.copy2(self.source, source)
            workbook = load_workbook(source)
            workbook["风险登记册"]["F4"] = period
            workbook.save(source)
            out = self.root / f"out-{len(period)}"
            with self.subTest(export_period=period), self.assertRaises((ValidationError, ValueError)):
                export_workbook(source, out)
            self.assertFalse((self.root / "escaped_period").exists())
            self.assertFalse((self.root / "absolute").exists())

    def test_preview_commit_token_rejects_source_or_decision_change(self) -> None:
        decision = create_decision()
        findings = (finding("F-create"),)
        preview = preview_changes(self.source, (decision,), findings)
        self.assertRegex(preview["commit_token"], r"^[0-9a-f]{64}$")
        changed = load_workbook(self.source)
        changed["风险登记册"]["B4"] = "只用于令牌失效的合成改动"
        changed.save(self.source)
        with self.assertRaisesRegex(ValidationError, "PREVIEW_STALE"):
            write_versioned_workbook(self.source, (decision,), findings, expected_commit_token=preview["commit_token"],
                                     timestamp="20260903_1212", output_dir=self.root / "versions")
        self.assertFalse((self.root / "versions" / "audit_risk_register_20260903_1212.xlsx").exists())

        shutil.copy2(SYNTHETIC_SOURCE, self.source)
        self.source_hash = sha256(self.source)
        preview = preview_changes(self.source, (decision,), findings)
        decision.name = "令牌之后被篡改的合成风险"
        with self.assertRaisesRegex(ValidationError, "PREVIEW_STALE"):
            write_versioned_workbook(self.source, (decision,), findings, expected_commit_token=preview["commit_token"],
                                     timestamp="20260903_1213", output_dir=self.root / "versions")
        self.assertFalse((self.root / "versions" / "audit_risk_register_20260903_1213.xlsx").exists())
        self.assert_source_unchanged()

    def test_build_failure_uses_private_temp_paths_and_publishes_nothing(self) -> None:
        decision, findings = create_decision(), (finding("F-create"),)
        token = preview_changes(self.source, (decision,), findings)["commit_token"]
        seen = []

        def fail_export(path, out_dir):
            seen.extend((Path(path), Path(out_dir)))
            raise RuntimeError("synthetic interruption")

        with patch("desktop.workbook_writer.export_workbook", side_effect=fail_export), self.assertRaises(RuntimeError):
            write_versioned_workbook(self.source, (decision,), findings, expected_commit_token=token,
                                     timestamp="20260903_1214", output_dir=self.root / "versions")
        final = self.root / "versions" / "audit_risk_register_20260903_1214.xlsx"
        self.assertTrue(seen)
        self.assertNotEqual(seen[0], final)
        self.assertFalse(final.exists())
        self.assertFalse(final.with_name(final.stem + "_data_export").exists())
        self.assert_source_unchanged()

    def test_external_export_race_is_not_overwritten_or_deleted(self) -> None:
        decision, findings = create_decision(), (finding("F-create"),)
        token = preview_changes(self.source, (decision,), findings)["commit_token"]
        destination = self.root / "versions"
        final_export = destination / "audit_risk_register_20260903_1215_data_export"
        real_export = export_workbook

        def external_race(path, out_dir):
            result = real_export(path, out_dir)
            final_export.mkdir(parents=True)
            (final_export / "external.marker").write_text("preserve", encoding="utf-8")
            return result

        with patch("desktop.workbook_writer.export_workbook", side_effect=external_race), self.assertRaisesRegex(ValidationError, "OUTPUT_EXISTS"):
            write_versioned_workbook(self.source, (decision,), findings, expected_commit_token=token,
                                     timestamp="20260903_1215", output_dir=destination)
        self.assertTrue((final_export / "external.marker").is_file())
        self.assertFalse((destination / "audit_risk_register_20260903_1215.xlsx").exists())
        self.assert_source_unchanged()

    def test_create_id_must_be_the_next_monotonic_risk_id(self) -> None:
        with self.assertRaises((ValidationError, ValueError)):
            preview_changes(self.source, (create_decision("F-create", "R999"),), (finding("F-create"),))
        preview = preview_changes(self.source, (create_decision("F-create", "R025"),), (finding("F-create"),))
        self.assertEqual(preview["new_risks"][0]["risk_id"], "R025")

    def test_unaccepted_unknown_duplicate_or_omitted_findings_are_atomic(self) -> None:
        cases = [
            ((create_decision(),), (finding("F-create", "待确认"),)),
            ((create_decision(),), (finding("other"),)),
            ((create_decision("F-create"), merge_decision("F-create")), (finding("F-create"),)),
            ((create_decision(),), (finding("F-create"), finding("omitted"))),
            ((create_decision(),), (finding("F-create"), finding("F-create"))),
        ]
        for decisions, findings in cases:
            with self.subTest(decisions=decisions):
                with self.assertRaises((ValidationError, ValueError)):
                    write_versioned_workbook(self.source, decisions, findings,
                                             expected_commit_token="invalid", timestamp="20260903_1201", output_dir=self.root / "versions")
                self.assertFalse((self.root / "versions" / "audit_risk_register_20260903_1201.xlsx").exists())
                self.assert_source_unchanged()

    def test_excluded_controls_are_not_written_and_merge_replaces_controls(self) -> None:
        result = self.commit(
            self.source, (merge_decision(), RiskDecision(action="exclude", finding_ids=("F-exclude",))),
            (finding("F-merge"), finding("F-exclude", "已排除")), timestamp="20260903_1202",
            output_dir=self.root / "versions",
        )
        wb = load_workbook(result.workbook_path, data_only=False)
        controls = wb["控制措施表"]
        matching = [tuple(controls.cell(row, col).value for col in range(1, 7))
                    for row in range(4, 154)
                    if controls.cell(row, 2).value == "R001" and controls.cell(row, 3).value == "2026H1"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0][3:], ("更新控制", 5, "是"))
        self.assertNotIn("F-exclude", str(matching))
        self.assert_source_unchanged()

    def test_empty_confirmed_controls_removes_current_controls(self) -> None:
        result = self.commit(
            self.source, (merge_decision(controls=()),), (finding("F-merge"),),
            timestamp="20260903_1203", output_dir=self.root / "versions",
        )
        wb = load_workbook(result.workbook_path, data_only=False)
        controls = wb["控制措施表"]
        self.assertFalse(any(controls.cell(row, 2).value == "R001" and controls.cell(row, 3).value == "2026H1"
                             for row in range(4, 154)))

    def test_merge_control_replacement_never_reuses_removed_control_id(self) -> None:
        decision = RiskDecision(
            action="merge", finding_ids=("F-merge",), risk_id="R024", name="更新后的合成风险",
            domain="采购与外包", description="更新风险描述", owner_dept="审计部",
            period="2025H2", likelihood=5, impact_scores={dim: 4 for dim in DIMS},
            rationale="更新后的合成依据", controls=(ConfirmedControl("替换后的控制", 5, True),),
        )
        result = self.commit(self.source, (decision,), (finding("F-merge"),),
                                          timestamp="20260903_1210", output_dir=self.root / "versions")
        controls = load_workbook(result.workbook_path, data_only=False)["控制措施表"]
        replacement_ids = [controls.cell(row, 1).value for row in range(4, 154)
                           if controls.cell(row, 2).value == "R024" and controls.cell(row, 3).value == "2025H2"]
        self.assertEqual(replacement_ids, ["C064"])
        self.assert_source_unchanged()

    def test_input_writes_preserve_formula_template_and_config_export_parity(self) -> None:
        result = self.commit(self.source, (create_decision(),), (finding("F-create"),),
                                          timestamp="20260903_1204", output_dir=self.root / "versions")
        source_wb = load_workbook(self.source, data_only=False)
        written = load_workbook(result.workbook_path, data_only=False)
        register = written["风险登记册"]
        row = next(row for row in range(4, 204) if register.cell(row, 1).value == "R025")
        for col in range(16, 26):
            self.assertEqual(register.cell(row, col).value, source_wb["风险登记册"].cell(row, col).value)
        self.assertEqual(register.cell(row, 26).value, "确认后的合成依据")
        old_export = self.root / "old_export"
        new_export = self.root / "new_export"
        export_workbook(self.source, old_export)
        export_workbook(result.workbook_path, new_export)
        self.assertEqual(json.loads((old_export / "config.json").read_text(encoding="utf-8")),
                         json.loads((new_export / "config.json").read_text(encoding="utf-8")))

    def test_safe_failures_leave_no_partial_files(self) -> None:
        collisions = self.root / "versions"
        collisions.mkdir()
        (collisions / "audit_risk_register_20260903_1205.xlsx").write_bytes(b"keep")
        with self.assertRaises(ValueError):
            self.commit(self.source, (create_decision(),), (finding("F-create"),),
                        timestamp="20260903_1205", output_dir=collisions)
        self.assertEqual((collisions / "audit_risk_register_20260903_1205.xlsx").read_bytes(), b"keep")
        self.assert_source_unchanged()
        for mutate in (self._remove_sheet, self._duplicate_risk_id, self._fill_capacity):
            source = self.root / f"{mutate.__name__}.xlsx"
            shutil.copy2(self.source, source)
            mutate(source)
            with self.subTest(mutate=mutate.__name__), self.assertRaises((ValidationError, ValueError)):
                write_versioned_workbook(source, (create_decision(),), (finding("F-create"),),
                                         expected_commit_token="invalid", timestamp="20260903_1206", output_dir=self.root / "failures")
            self.assertFalse((self.root / "failures" / "audit_risk_register_20260903_1206.xlsx").exists())

    def test_same_output_name_is_exclusively_created_by_one_concurrent_writer(self) -> None:
        destination = self.root / "versions"
        start = threading.Barrier(2)

        def attempt():
            start.wait(timeout=5)
            try:
                return self.commit(
                    self.source, (create_decision(),), (finding("F-create"),),
                    timestamp="20260903_1211", output_dir=destination,
                )
            except Exception as exc:  # The public API intentionally returns safe errors.
                return exc

        with ThreadPoolExecutor(max_workers=2) as pool:
            outcomes = list(pool.map(lambda _: attempt(), range(2)))
        winners = [outcome for outcome in outcomes if isinstance(outcome, WorkbookWriteResult)]
        losers = [outcome for outcome in outcomes if isinstance(outcome, Exception)]
        self.assertEqual(len(winners), 1)
        self.assertEqual(len(losers), 1)
        self.assertEqual(str(losers[0]), "OUTPUT_EXISTS")
        self.assertTrue(winners[0].workbook_path.is_file())
        self.assertGreater(winners[0].workbook_path.stat().st_size, 0)
        self.assertIn("风险登记册", load_workbook(winners[0].workbook_path, data_only=False).sheetnames)
        self.assert_source_unchanged()

    def test_merge_requires_existing_exact_risk_period_and_create_id_is_valid_unused(self) -> None:
        missing = merge_decision()
        missing.period = "2099H1"
        with self.assertRaises((ValidationError, ValueError)):
            write_versioned_workbook(self.source, (missing,), (finding("F-merge"),),
                                     expected_commit_token="invalid", timestamp="20260903_1207", output_dir=self.root / "versions")
        with self.assertRaises((ValidationError, ValueError)):
            write_versioned_workbook(self.source, (create_decision("F-create", "bad"),), (finding("F-create"),),
                                     expected_commit_token="invalid", timestamp="20260903_1208", output_dir=self.root / "versions")
        with self.assertRaises((ValidationError, ValueError)):
            write_versioned_workbook(self.source, (create_decision("F-create", "R001"),), (finding("F-create"),),
                                     expected_commit_token="invalid", timestamp="20260903_1208", output_dir=self.root / "versions")

    def test_exporter_callable_matches_legacy_content_and_cli_failure_is_chinese(self) -> None:
        result = export_workbook(self.source, self.root / "export")
        self.assertEqual(result["risks"], 48)
        self.assertEqual(result["controls"], 63)
        self.assertEqual(result["periods"], ["2025H2", "2026H1"])
        self.assertTrue((self.root / "export" / "2026H1" / "risks.csv").is_file())
        bad = self.root / "bad.xlsx"
        shutil.copy2(self.source, bad)
        wb = load_workbook(bad)
        wb["参数配置"]["B3"] = 0.1
        wb.save(bad)
        process = subprocess.run([sys.executable, "tools/export_from_excel.py", "--xlsx", str(bad)],
                                 cwd=ROOT, capture_output=True, text=True)
        self.assertNotEqual(process.returncode, 0)
        self.assertIn("[错误] 权重行", process.stdout + process.stderr)

    @staticmethod
    def _remove_sheet(path: Path) -> None:
        wb = load_workbook(path)
        del wb["风险登记册"]
        wb.save(path)

    @staticmethod
    def _duplicate_risk_id(path: Path) -> None:
        wb = load_workbook(path)
        ws = wb["风险登记册"]
        ws["A30"] = ws["A4"].value
        ws["F30"] = ws["F4"].value
        wb.save(path)

    @staticmethod
    def _fill_capacity(path: Path) -> None:
        wb = load_workbook(path)
        ws = wb["风险登记册"]
        for row in range(4, 204):
            if not ws.cell(row, 1).value:
                ws.cell(row, 1).value = f"R{row + 1000:03d}"
        wb.save(path)

    @staticmethod
    def _all_assessed(export_dir: Path, periods: list[str]) -> list[dict]:
        assessed = []
        for period in periods:
            config, risks, controls = load_dataset(export_dir / period, export_dir / "config.json")
            assessed.extend(assess_all(risks, controls, config))
        return assessed


if __name__ == "__main__":
    unittest.main()
