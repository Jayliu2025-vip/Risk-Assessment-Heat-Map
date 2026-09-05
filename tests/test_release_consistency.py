# -*- coding: utf-8 -*-
"""v1.2 发布物口径与网页默认配置一致性测试。"""

import json
import math
import re
import shutil
import subprocess
import unittest
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from tools import common
from tools.export_from_excel import read_config


ROOT = Path(__file__).resolve().parents[1]


def read_text(relative_path):
    return (ROOT / relative_path).read_text(encoding="utf-8")


class ReleaseConsistencyTests(unittest.TestCase):
    def test_scoring_anchor_release_artifacts_cannot_drift(self):
        canonical = json.loads(read_text("data/scoring_anchors.json"))
        self.assertEqual([group["key"] for group in canonical], ["likelihood", *common.DIMS])

        script = read_text("web/scoring_anchors.js").strip()
        prefix = "window.SCORING_ANCHORS = "
        self.assertTrue(script.startswith(prefix))
        self.assertEqual(json.loads(script[len(prefix):].removesuffix(";")), canonical)

        build_source = read_text("tools/build_excel.py")
        self.assertIn("load_scoring_anchors", build_source)
        self.assertNotIn("blocks = [", build_source)

        html = read_text("web/risk_heatmap.html")
        self.assertIn('<script src="sample_data.js"></script>', html)
        self.assertIn('<script src="scoring_anchors.js"></script>', html)
        self.assertIn("SCORING_ANCHORS.map", html)
        self.assertNotIn("const ANCHORS = [", html)

    def test_release_artifacts_share_v12_application_version(self):
        expected_version = "1.2"

        self.assertEqual(
            getattr(common, "MODEL_VERSION", None),
            expected_version,
            "tools.common.MODEL_VERSION 必须声明当前发布版本",
        )
        self.assertEqual(common.DEFAULT_CONFIG["version"], expected_version)
        self.assertEqual(
            json.loads(read_text("data/export/config.json"))["version"],
            expected_version,
        )

        sample_script = read_text("web/sample_data.js").strip()
        sample_prefix = "window.SAMPLE_DATA = "
        self.assertTrue(sample_script.startswith(sample_prefix))
        sample_data = json.loads(
            sample_script[len(sample_prefix):].removesuffix(";")
        )
        self.assertEqual(sample_data["config"]["version"], expected_version)

        self.assertIn("当前版本：`1.2`", read_text("README.md"))
        manual = read_text("docs/使用手册.md")
        self.assertIn("> 版本 1.2", manual)
        self.assertIn('"version": "1.2"', manual)

        html = read_text("web/risk_heatmap.html")
        self.assertIn('const APP_VERSION = "1.2";', html)
        self.assertRegex(html, r"version\s*:\s*APP_VERSION")
        self.assertIn('id="app-version"', html)
        self.assertIn('$("app-version").textContent=`v${APP_VERSION}`', html)

        workbook = load_workbook(ROOT / "audit_risk_register.xlsx", data_only=False)
        self.assertIn("v1.2", workbook["使用说明"]["A1"].value)
        self.assertEqual(
            read_config(workbook["参数配置"])["version"], expected_version
        )

    def test_current_user_copy_does_not_contain_legacy_wording(self):
        forbidden_by_file = {
            "README.md": [
                "21 项", "6 领域 16 风险", "24 项评分模型一致性测试",
                "Σ(wᵢ×维度ᵢ), floor × 最高维度分",
            ],
            "docs/使用手册.md": [
                "五维", "六项打分", "16 条", "6 个领域", "21 项",
                "24 项评分模型一致性测试", "B3:B7", "B12:B15", "A25:G38",
                "G 列变红", "I2:I8", "登记册 Q 列", "登记册 M~R 列",
                "A~L 列", "G~L 打分", "登记册 M 列", "$A$26:$F$38",
                "五个权重输入框", "参数配置页 B8 合计格", "把五个权重",
                "30%+25%+15%+15%+15%",
            ],
            "tools/build_excel.py": ["五维", "六项打分", "B12:B15"],
            "tools/common.py": ["Σ(wᵢ×维度ᵢ), floor × MAX(各维度)"],
            "tools/generate_report.py": ["五维加权"],
            "tools/sample_data.py": ["6 领域 × 16 项风险"],
            "tests/test_scoring.py": ["22.00 → 18.70", "12.00 → 10.20"],
            "web/risk_heatmap.html": ["五维影响权重", "五维加权"],
        }
        for relative_path, forbidden_terms in forbidden_by_file.items():
            content = read_text(relative_path)
            for term in forbidden_terms:
                with self.subTest(file=relative_path, term=term):
                    self.assertNotIn(term, content)

    def test_release_documentation_uses_current_ranges_and_dimensions(self):
        readme = read_text("README.md")
        manual = read_text("docs/使用手册.md")
        sample_source = read_text("tools/sample_data.py")

        for dimension in (
            "经济损失",
            "合规法律",
            "运营中断",
            "声誉舆情",
            "舞弊风险",
            "战略影响",
            "数据安全",
            "健康安全",
        ):
            with self.subTest(dimension=dimension):
                self.assertIn(dimension, readme)
        self.assertIn("4 大类 12 领域", readme)
        self.assertIn("每期 24 条", sample_source)
        self.assertIn("4 大类 12 领域", sample_source)
        for expected in (
            "B3:B10",
            "B14:B17",
            "A25:J38",
            "J 列变红",
            "I1:I8",
            "P~Z",
            "登记册 T 列",
            "登记册 P~U 列",
            "$A$26:$I$38",
            "24 项评分模型测试 + 发布一致性测试",
        ):
            with self.subTest(expected=expected):
                self.assertIn(expected, manual)
        self.assertIn("262 项 Python 测试", readme)
        self.assertIn("13 项 Playwright 流程", readme)
        self.assertIn("Σ(wᵢ×维度ᵢ)/Σ(已评分维度的wᵢ)", readme)
        self.assertIn("八个权重输入框", manual)
        self.assertIn("参数配置页 B11 合计格", manual)
        self.assertIn("25%+20%+12%+10%+10%+10%+8%+5%", manual)
        self.assertIn(
            "| 领域 | 经济 | 合规 | 运营 | 声誉 | 舞弊 | 战略 | 数据 | 健康安全 |",
            manual,
        )

    @unittest.skipUnless(
        shutil.which("node"),
        "需要 Node.js 才能运行网页配置行为测试；Node 不是本项目运行依赖。",
    )
    def test_normalize_config_migrates_legacy_rows_and_rejects_bad_fields(self):
        html = read_text("web/risk_heatmap.html")
        config_code = html[html.index("const DIMS ="):html.index("const LS_KEY")]
        script = config_code + r'''
const migrated=normalizeConfig({
  weights:{imp_financial:.30,imp_compliance:.25,imp_operation:.15,imp_reputation:.15,imp_fraud:.15},
  domain_weights:{Legacy:{imp_financial:.40,imp_compliance:.25,imp_operation:.10,imp_reputation:.15,imp_fraud:.10}},
  impact_floor_factor:"not-a-number",
  reduction_map:{"2":Infinity,"3":.5},
  thresholds:{extreme:NaN,high:13},
  ref_reduction:Infinity,
  unknown:"must-not-leak"
});
console.log(JSON.stringify(migrated));
'''
        completed = subprocess.run(
            ["node", "-"], input=script, text=True, capture_output=True, check=True
        )
        config = json.loads(completed.stdout)
        canonical_fields = {
            "version", "weights", "domain_weights", "impact_floor_factor",
            "reduction_map", "thresholds", "ref_reduction",
        }
        dimensions = {
            "imp_financial", "imp_compliance", "imp_operation", "imp_reputation",
            "imp_fraud", "imp_strategy", "imp_data", "imp_hse",
        }
        self.assertEqual(set(config), canonical_fields)
        self.assertEqual(config["version"], "1.2")
        self.assertEqual(config["impact_floor_factor"], 0.75)
        self.assertEqual(config["ref_reduction"], 0.4)
        self.assertEqual(config["reduction_map"]["2"], 0.15)
        self.assertEqual(config["thresholds"]["extreme"], 20)
        for row in [config["weights"], *config["domain_weights"].values()]:
            self.assertEqual(set(row), dimensions)
            self.assertAlmostEqual(sum(row.values()), 1.0)
        self.assertAlmostEqual(config["weights"]["imp_strategy"], 0.10)
        self.assertAlmostEqual(config["weights"]["imp_data"], 0.08)
        self.assertAlmostEqual(config["weights"]["imp_hse"], 0.05)

    def test_import_config_replaces_previous_normalized_config(self):
        html = read_text("web/risk_heatmap.html")
        body = html.split("function importConfig(file){", 1)[1].split(
            "function csvEscape", 1
        )[0]
        self.assertIn("state.config=normalizeConfig(j)", body)
        self.assertNotIn("domainWeights", body)

    def test_edit_risk_restores_likelihood_input(self):
        html = read_text("web/risk_heatmap.html")
        body = html.split("function editRisk(rid){", 1)[1].split(
            "function delRisk", 1
        )[0]
        self.assertIn('$("f-lik").value=r.likelihood', body)

    def test_workbook_layout_and_period_dropdowns_match_current_model(self):
        workbook = load_workbook(ROOT / "audit_risk_register.xlsx", data_only=False)
        config = workbook["参数配置"]
        self.assertEqual(config["I25"].value, "健康安全")
        self.assertEqual(config["J25"].value, "行合计")
        self.assertEqual(
            [config.cell(row=row, column=9).value for row in range(1, 5)],
            ["评估期间清单", "全部", "2025H2", "2026H1"],
        )

        register = workbook["风险登记册"]
        expected_headers = [
            "综合影响",
            "固有风险",
            "关键控制分",
            "折减系数",
            "剩余风险",
            "剩余等级",
            "影响档\n(辅助)",
            "固有点\n(辅助)",
            "建议审计频次",
            "控制挽回率",
            "打分依据",
        ]
        self.assertEqual(
            [register.cell(row=3, column=column).value for column in range(16, 27)],
            expected_headers,
        )

        dropdown_formulas = {
            sheet_name: {
                validation.formula1
                for validation in workbook[sheet_name].data_validations.dataValidation
            }
            for sheet_name in ("风险登记册", "控制措施表", "热力图")
        }
        self.assertIn("=参数配置!$I$3:$I$8", dropdown_formulas["风险登记册"])
        self.assertIn("=参数配置!$I$3:$I$8", dropdown_formulas["控制措施表"])
        self.assertIn("=参数配置!$I$2:$I$8", dropdown_formulas["热力图"])

    def test_workbook_recovery_formula_guards_blank_and_zero_inherent_risk(self):
        workbook = load_workbook(ROOT / "audit_risk_register.xlsx", data_only=False)
        self.assertEqual(
            workbook["风险登记册"]["Y4"].value,
            '=IF($A4="","",IF(OR(Q4="",Q4=0),"",(Q4-T4)/Q4))',
        )

    def test_workbook_priority_ranking_uses_residual_score(self):
        workbook = load_workbook(ROOT / "audit_risk_register.xlsx", data_only=False)
        register = workbook["风险登记册"]
        summary = workbook["汇总与优先级"]
        self.assertEqual(register["AA3"].value, "优先级排序键\n(辅助)")
        self.assertTrue(register.column_dimensions["AA"].hidden)
        key_formula = register["AA4"].value
        self.assertIn("T4", key_formula)
        self.assertIn("ROW()", key_formula)
        self.assertEqual(register["AB3"].value, "选定期间排序键\n(辅助)")
        self.assertTrue(register.column_dimensions["AB"].hidden)
        period_key_formula = register["AB4"].value
        self.assertIn("'汇总与优先级'!$B$2", period_key_formula)
        self.assertIn("$F4", period_key_formula)
        self.assertIn("$AA4", period_key_formula)
        self.assertIn("NA()", period_key_formula)

        selector_formula = summary["L21"].value
        self.assertIn("AGGREGATE(14,6", selector_formula)
        self.assertIn("风险登记册!$AB$4:$AB$203", selector_formula)
        self.assertNotIn("风险登记册!$Q$4:$Q$203", selector_formula)

    def test_workbook_priority_list_selects_period_top_rows_via_index_match(self):
        workbook = load_workbook(ROOT / "audit_risk_register.xlsx", data_only=False)
        summary = workbook["汇总与优先级"]
        self.assertTrue(summary.column_dimensions["L"].hidden)
        selector_formula = summary["L21"].value
        self.assertNotIn("/((", selector_formula)
        self.assertNotIn("风险登记册!$F$4:$F$203", selector_formula)
        self.assertIn("ROWS($L$21:L21)", selector_formula)

        self.assertEqual(
            summary["A21"].value,
            '=IF($L21="","",ROWS($A$21:A21))',
        )
        for column in range(2, 11):
            formula = summary.cell(row=21, column=column).value
            with self.subTest(column=column):
                self.assertIn("INDEX(", formula)
                self.assertIn("MATCH($L21,风险登记册!$AA$4:$AA$203,0)", formula)
                self.assertNotIn("风险登记册!A4", formula)
        self.assertTrue(summary["K21"].value.startswith('=IF($L21="","",'))
        self.assertNotIn("风险登记册!", summary["K21"].value)

    def test_report_bubble_offsets_keep_eight_risks_unique_inside_cell(self):
        from tools import generate_report

        offset_factory = getattr(generate_report, "bubble_offsets", None)
        self.assertTrue(
            callable(offset_factory),
            "报告生成器必须提供可测试的 bubble_offsets(count)",
        )
        offsets = offset_factory(8)
        self.assertEqual(len(offsets), 8)
        self.assertEqual(len(set(offsets)), 8)
        self.assertTrue(all(abs(x) <= 0.5 and abs(y) <= 0.5 for x, y in offsets))
        minimum_spacing = min(
            math.hypot(ax - bx, ay - by)
            for index, (ax, ay) in enumerate(offsets)
            for bx, by in offsets[index + 1:]
        )
        self.assertGreaterEqual(minimum_spacing, 0.20)

        cfg, risks, controls = common.load_dataset(ROOT / "data/export/2026H1")
        assessed = common.assess_all(risks, controls, cfg)
        cell_counts = Counter(
            (risk["likelihood"], generate_report.impact_cell(risk["impact"]))
            for risk in assessed
        )
        self.assertEqual(max(cell_counts.values()), 8)
        for count in cell_counts.values():
            positions = offset_factory(count)
            self.assertEqual(len(positions), len(set(positions)))

    @unittest.skipUnless(
        shutil.which("node"),
        "需要 Node.js 才能运行跨端影响档行为测试；Node 不是本项目运行依赖。",
    )
    def test_python_and_web_impact_cells_use_half_up_rounding(self):
        from tools import generate_report

        impact_cell = getattr(generate_report, "impact_cell", None)
        self.assertTrue(callable(impact_cell), "报告生成器必须提供 impact_cell(value)")
        values = [2.5, 3.5, 4.5]
        expected = [3, 4, 5]
        self.assertEqual([impact_cell(value) for value in values], expected)

        html = read_text("web/risk_heatmap.html")
        match = re.search(
            r"function impactCell\(value\)\{.*?\n\}", html, re.DOTALL
        )
        self.assertIsNotNone(match, "网页必须提供 impactCell(value)")
        script = (
            match.group(0)
            + "\nconsole.log(JSON.stringify([2.5,3.5,4.5].map(impactCell)));"
        )
        completed = subprocess.run(
            ["node", "-"], input=script, text=True, capture_output=True, check=True
        )
        self.assertEqual(json.loads(completed.stdout), expected)

    @unittest.skipUnless(
        shutil.which("node"),
        "需要 Node.js 才能运行网页气泡定位行为测试；Node 不是本项目运行依赖。",
    )
    def test_web_heatmap_offsets_support_eight_risks_around_cell_center(self):
        html = read_text("web/risk_heatmap.html")
        match = re.search(
            r"function cellOffsets\(count\)\{.*?\n\}", html, re.DOTALL
        )
        self.assertIsNotNone(match, "网页必须提供 cellOffsets(count)")
        script = match.group(0) + "\nconsole.log(JSON.stringify(cellOffsets(8)));"
        completed = subprocess.run(
            ["node", "-"], input=script, text=True, capture_output=True, check=True
        )
        offsets = json.loads(completed.stdout)
        self.assertEqual(len(offsets), 8)
        self.assertEqual(len({tuple(point) for point in offsets}), 8)
        self.assertTrue(
            all(abs(x) <= 0.5 and abs(y) <= 0.5 for x, y in offsets)
        )

        draw_body = html.split("function drawHeat(elId,rows,mode){", 1)[1].split(
            "function renderTrend", 1
        )[0]
        self.assertIn("impactCell(a.impact)", draw_body)
        self.assertNotIn("Math.round(a.impact)", draw_body)
        self.assertIn("cellOffsets(items.length)", draw_body)
        self.assertRegex(draw_body, r"ii\s*\+\s*dx")
        self.assertRegex(draw_body, r"li\s*\+\s*dy")
        self.assertNotIn("a.impact+dx", draw_body)
        self.assertNotRegex(draw_body, r"%\s*OFFS\.length")

    def test_heatmap_copy_distinguishes_exact_impact_from_rounded_cell(self):
        from tools import generate_report

        figure, axis = generate_report.plt.subplots()
        try:
            generate_report.draw_bubble_heatmap(
                axis, [], common.DEFAULT_CONFIG, "inherent", "测试期"
            )
            self.assertEqual(axis.get_xlabel(), "综合影响档（四舍五入）")
        finally:
            generate_report.plt.close(figure)

        html = read_text("web/risk_heatmap.html")
        self.assertIn("综合影响档（四舍五入）", html)
        self.assertIn("四舍五入影响档", html)
        self.assertIn(
            "精确影响 ${a.impact.toFixed(2)} / 影响档 ${ii}", html
        )

        manual = read_text("docs/使用手册.md")
        self.assertIn("横轴=综合影响档（精确综合影响四舍五入）", manual)
        self.assertNotIn("横轴=综合影响，", manual)

    def test_web_default_config_has_all_eight_weights_summing_to_one(self):
        html = read_text("web/risk_heatmap.html")
        match = re.search(
            r"const\s+DEFAULT_CFG\s*=\s*\{.*?weights\s*:\s*\{([^}]*)\}",
            html,
            re.DOTALL,
        )
        self.assertIsNotNone(match, "未找到 DEFAULT_CFG.weights")
        weights = {
            key: float(value)
            for key, value in re.findall(
                r"(imp_[a-z]+)\s*:\s*(\d*\.?\d+)", match.group(1)
            )
        }
        expected_dimensions = {
            "imp_financial",
            "imp_compliance",
            "imp_operation",
            "imp_reputation",
            "imp_fraud",
            "imp_strategy",
            "imp_data",
            "imp_hse",
        }
        self.assertEqual(set(weights), expected_dimensions)
        self.assertAlmostEqual(sum(weights.values()), 1.0)

    def test_web_uses_canonical_impact_floor_factor(self):
        html = read_text("web/risk_heatmap.html")
        self.assertRegex(html, r"impact_floor_factor\s*:\s*\.75")
        self.assertRegex(html, r"cfg\.impact_floor_factor")

    def test_embedded_sample_config_matches_exported_config(self):
        exported = json.loads(read_text("data/export/config.json"))
        script = read_text("web/sample_data.js").strip()
        prefix = "window.SAMPLE_DATA = "
        self.assertTrue(script.startswith(prefix))
        embedded = json.loads(script[len(prefix):].removesuffix(";"))
        self.assertEqual(embedded["config"], exported)


if __name__ == "__main__":
    unittest.main()
